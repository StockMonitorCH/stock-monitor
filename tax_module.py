#!/usr/bin/env python3
"""
tax_module.py – Steuerauszug-Modul für Stock Monitor v5.0
==========================================================
Dieses Modul ist bewusst als eigenständige Datei ausgelagert,
damit Steuerregeln (die sich jährlich ändern) unabhängig vom
Haupt-Programm aktualisiert werden können.

Unterstützte Länder:
  AT – Österreich  (KESt 27,5 % flat auf Dividenden und Kursgewinne)
  CH – Schweiz     (Verrechnungssteuer + Einkommenssteuer auf Dividenden)
  DE – Deutschland (Abgeltungssteuer 25 % + Soli)
  UK – Grossbritannien (Income Tax auf Dividenden, CGT auf Kursgewinne)
  US – USA         (Federal withholding tax, Capital Gains)

Haftungsausschluss:
  Bitte verwenden Sie den originalen Bankauszug für Ihre Steuererklärung.
  Dieses Portfolio kann bis ca. 2 % Abweichung aufweisen (Banklogik).
  Gebühren, Lombardkredite u. ä. werden nicht berücksichtigt.

Version: 2025-01  (jährlich aktualisierbar)
"""

import os
from tax_translations import TRT, get_tax_language

# ── Steuersätze & Regeln – NUR HIER ÄNDERN ─────────────────────────────────
TAX_RULES = {
    "AT": {
        "name":        "Österreich",
        "flag":        "🇦🇹",
        "year":        2025,
        # KESt 27,5 % flat auf Dividenden und realisierte Kursgewinne
        "kest_rate":        0.275,
        # Für Einkommen unter Grenzsteuersatz 27,5 %: Regelbesteuerungsoption möglich
        # (Antrag beim Finanzamt) – hier nicht berechnet, da individuell
        "sparpaket_note":   "Regelbesteuerungsoption möglich (§ 27a Abs. 5 EStG)",
        # Verlustausgleich: realisierte Verluste können mit Gewinnen/Dividenden verrechnet werden
        "verlustausgleich":  True,
        "verlustausgleich_note": "Verlustausgleich innerhalb Depot möglich (Depotbank führt automatisch durch)",
        "disclaimer":  (
            "AT: KESt 27,5 % wird von der Depotbank automatisch abgeführt (bei österr. Depot). "
            "Ausländische Dividenden: Quellensteuer je nach DBA, anrechenbar auf KESt. "
            "Verlustausgleich innerhalb des Depots erfolgt automatisch durch die Depotbank. "
            "Regelbesteuerungsoption (§ 27a Abs. 5 EStG) nur auf Antrag beim Finanzamt."
        ),
    },
    "CH": {
        "name":        "Schweiz",
        "flag":        "🇨🇭",
        "year":        2025,
        # Verrechnungssteuer (35 %) auf Schweizer Dividenden → rückforderbar
        "vst_rate":    0.35,
        # Einkommenssteuer auf ausländische Dividenden (Schätzung Bund + Kanton)
        "div_tax_foreign": 0.30,
        # Kapitalgewinnsteuer Privatperson = 0 (Bundesebene)
        "cgt_rate":    0.00,
        "cgt_note":    "Keine Kapitalgewinnsteuer für Privatpersonen (Bund)",
        "wealth_tax":  True,
        "wealth_tax_note": "Vermögenssteuer: Marktwert aller Positionen steuerpflichtig (kantonal)",
        "disclaimer":  (
            "CH: Verrechnungssteuer (35 %) auf Schweizer Dividenden ist rückforderbar "
            "(Wertschriftenverzeichnis DA-1 / R-US). Ausländische Dividenden: Quellensteuer "
            "je nach DBA. Kantonale Einkommenssteuern variieren stark."
        ),
    },
    "DE": {
        "name":        "Deutschland",
        "flag":        "🇩🇪",
        "year":        2025,
        # Abgeltungssteuer 25 % + 5,5 % Soli = 26,375 %
        "withholding_rate": 0.25,
        "soli_rate":   0.055,   # 5,5 % auf Abgeltungssteuer
        "effective_rate": 0.26375,
        "freistellungsauftrag": 1000,   # EUR/Person (ab 2023)
        "cgt_rate":    0.26375,
        "cgt_note":    "Abgeltungssteuer 25 % + Soli 5,5 % auf Abgeltungsteuer",
        "disclaimer":  (
            "DE: Abgeltungssteuer 25 % zzgl. 5,5 % Solidaritätszuschlag. "
            "Freistellungsauftrag 1.000 € p.P. (2025). Kirchensteuer ggf. zusätzlich. "
            "Verlustverrechnung über Steuerbescheinigung der Depotbank."
        ),
    },
    "UK": {
        "name":        "Grossbritannien",
        "flag":        "🇬🇧",
        "year":        2024,   # UK Steuerjahr: April–April
        # Dividend Allowance £500 (ab April 2024)
        "dividend_allowance": 500,
        # Dividendensteuer (Basic Rate)
        "div_basic_rate":  0.0875,
        "div_higher_rate": 0.3375,
        "div_additional_rate": 0.3935,
        # Capital Gains Tax: Allowance £3.000 (ab April 2024)
        "cgt_allowance": 3000,
        "cgt_basic_rate":  0.10,
        "cgt_higher_rate": 0.20,
        "cgt_note":        "Basic/Higher Rate abhängig vom Gesamteinkommen",
        "disclaimer":  (
            "UK: Dividend Allowance £500 p.a. (ab April 2024). "
            "CGT Annual Exempt Amount £3.000 (ab April 2024). "
            "Steuersatz abhängig vom Gesamteinkommen. Gilt für UK-Steuerresidenten."
        ),
    },
    "US": {
        "name":        "USA",
        "flag":        "🇺🇸",
        "year":        2025,
        # Quellensteuer auf Dividenden für Nicht-US-Personen (Standard)
        "nra_withholding": 0.30,
        # Mit DBA (z. B. CH-US): reduziert auf 15 %
        "dba_ch_rate":  0.15,
        "dba_de_rate":  0.15,
        # Qualified Dividends (US-Resident): 0/15/20 % je nach Einkommensstufe
        "qualified_div_rates": [0.00, 0.15, 0.20],
        # Long-term CGT (US-Resident): 0/15/20 %
        "cgt_longterm": [0.00, 0.15, 0.20],
        # Short-term CGT = ordentlicher Einkommenssteuersatz
        "cgt_shortterm_note": "Short-term CGT = ordentlicher Einkommenssteuersatz",
        "disclaimer":  (
            "US: Für Nicht-US-Personen gilt 30 % Quellensteuer auf Dividenden "
            "(DBA CH/DE: 15 %). Für US-Residents: Qualified Dividends 0/15/20 %, "
            "Long-term Capital Gains 0/15/20 % abhängig vom Einkommen."
        ),
        "disclaimer_en": (
            "US: For non-US persons, 30 % withholding tax applies to dividends "
            "(treaty CH/DE: 15 %). For US residents: Qualified Dividends 0/15/20 %, "
            "Long-term Capital Gains 0/15/20 % depending on income."
        ),
    },
}

def _disclaimer_global() -> str:
    """Gibt den globalen Disclaimer in der aktuellen Sprache zurück."""
    return "⚠️  " + TRT("tax_disclaimer_global")

# ── Zahlenformatierung (wird von _build_tax_dialog gesetzt) ──────────────────
_TAX_NUMBER_FORMAT = "CH"   # Voreinstellung

def _fmt(value, decimals=2):
    """Zahlenformatierung gemäss _TAX_NUMBER_FORMAT (CH / DE / US)."""
    formatted = f"{abs(value):,.{decimals}f}"   # Basis: 35,969.57
    if _TAX_NUMBER_FORMAT == "CH":
        formatted = formatted.replace(',', "'")                              # 35'969.57
    elif _TAX_NUMBER_FORMAT == "DE":
        formatted = formatted.replace(',', 'X').replace('.', ',').replace('X', '.')  # 35.969,57
    # US: unverändert                                                        # 35,969.57
    return f"-{formatted}" if value < 0 else formatted

# ── Berechnungsfunktionen ────────────────────────────────────────────────────

def _fetch_real_dividends(ticker_sym, year, positions_for_sym=None):
    """
    Holt tatsächlich bezahlte Dividenden für ein Symbol im angegebenen Jahr.
    Berücksichtigt den Ex-Dividend-Tag und den tatsächlichen Bestand an diesem Tag.

    positions_for_sym: Liste der Portfolio-Positionen für dieses Symbol,
                       jede mit 'buy_date', 'quantity' (und optional 'sell_date').
                       Wenn None → einfache Summe über qty (ungenau).

    Gibt (total_dividends_gesamt, div_detail_list) zurück.
    total_dividends_gesamt = Summe aller Dividenden × Bestand am Ex-Day
    """
    import yfinance as yf
    import threading as _threading
    from datetime import date as _date, datetime as _dt

    # ticker.history() mit actions=True statt ticker.dividends — selber Code-Pfad
    # wie die Chart-Worker (curl_cffi-sicher auf dem Qt-Haupt-Thread)
    _hist_result = [None]
    def _fetch_in_thread():
        try:
            t = yf.Ticker(ticker_sym)
            data = t.history(start=f"{year}-01-01", end=f"{year+1}-01-01",
                             interval="1d", actions=True)
            if data is not None and not data.empty and 'Dividends' in data.columns:
                _hist_result[0] = data['Dividends'][data['Dividends'] > 0]
            else:
                _hist_result[0] = None
        except Exception:
            _hist_result[0] = None
    _th = _threading.Thread(target=_fetch_in_thread, daemon=True)
    _th.start()
    _th.join(timeout=15)

    try:
        year_divs = _hist_result[0]
        if year_divs is None or (hasattr(year_divs, 'empty') and year_divs.empty):
            return 0.0, []

        total = 0.0
        detail = []

        for ex_timestamp, div_per_share in year_divs.items():
            # Ex-Dividend-Datum normalisieren
            try:
                ex_day = ex_timestamp.date() if hasattr(ex_timestamp, 'date') else \
                         _dt.strptime(str(ex_timestamp)[:10], "%Y-%m-%d").date()
            except Exception:
                continue

            # Bestand am Ex-Day berechnen
            if positions_for_sym:
                qty_at_ex = 0.0
                for pos in positions_for_sym:
                    try:
                        buy_d = _dt.strptime(
                            pos.get('buy_date','')[:10], "%Y-%m-%d").date()
                    except Exception:
                        buy_d = _date(1900, 1, 1)

                    # Verkaufsdatum: falls vorhanden
                    sell_d_raw = pos.get('sell_date', '')
                    try:
                        sell_d = _dt.strptime(
                            str(sell_d_raw)[:10], "%Y-%m-%d").date() \
                            if sell_d_raw else _date(9999, 12, 31)
                    except Exception:
                        sell_d = _date(9999, 12, 31)

                    # Position war am Ex-Day aktiv: buy_date <= ex_day < sell_date
                    if buy_d <= ex_day < sell_d:
                        qty_at_ex += float(pos.get('quantity', 0))
            else:
                # Fallback: aktueller Bestand (ungenau für unterjährige Käufe/Verkäufe)
                qty_at_ex = float(
                    sum(p.get('quantity', 0) for p in (positions_for_sym or [])))

            div_amount = qty_at_ex * float(div_per_share)
            total += div_amount
            detail.append({
                "ex_date":       str(ex_day),
                "div_per_share": float(div_per_share),
                "qty_at_ex":     qty_at_ex,
                "div_amount":    div_amount,
            })

        return total, detail

    except Exception:
        return 0.0, []


def calc_ch(positions, currency="CHF"):
    """
    CH-Steuerauszug:
    - Verrechnungssteuer (35 %) auf CH-Dividenden (rückforderbar)
    - Ausländische Quellensteuer (DBA-abhängig)
    - Vermögenssteuerwert (Marktwert)
    Dividenden: tatsächlich bezahlte Beträge aus yfinance (Steuerjahr lt. TAX_RULES)
    """
    import yfinance as yf
    rows = []
    total_dividends      = 0.0
    total_vst            = 0.0
    total_foreign_tax    = 0.0
    total_market_value   = 0.0
    rules = TAX_RULES["CH"]
    year  = rules["year"]

    for pos in positions:
        ticker   = pos.get("ticker", "")
        name     = pos.get("name", ticker)
        qty      = float(pos.get("quantity", 0))
        price    = float(pos.get("current_price", 0))
        country  = pos.get("country", "").upper()

        market_val = qty * price

        # Tatsächliche Dividenden: vorberechnet oder direkt holen
        if "_div_total" in pos:
            div_total = pos["_div_total"]
        else:
            raw = pos.get("raw_positions") or [{"quantity": qty, "buy_date": ""}]
            div_total, _ = _fetch_real_dividends(ticker, year, positions_for_sym=raw)

        # Ticker-Herkunft: CH = Verrechnungssteuer, sonst Quellensteuer
        # Erkennung: .SW Suffix oder explizites country=="CH"
        is_ch = country == "CH" or ticker.upper().endswith(".SW")
        if is_ch:
            vst         = div_total * rules["vst_rate"]
            foreign_tax = 0.0
            tax_note    = "VSt 35 % (rückforderbar)"
        elif country in ("US", ""):
            vst         = 0.0
            # US-DBA CH: 15 %
            foreign_tax = div_total * 0.15
            tax_note    = "US QSt 15 % (DBA CH-US)"
        elif country == "DE":
            vst         = 0.0
            foreign_tax = div_total * 0.15
            tax_note    = "DE QSt 15 % (DBA CH-DE)"
        elif country == "GB":
            vst         = 0.0
            foreign_tax = div_total * 0.00   # UK: kein Quellensteuerabzug für Schweizer
            tax_note    = "UK: keine Quellensteuer"
        elif country == "FR":
            vst         = 0.0
            foreign_tax = div_total * 0.128  # FR DBA CH: 12.8 %
            tax_note    = "FR QSt 12.8 % (DBA CH-FR)"
        else:
            vst         = 0.0
            foreign_tax = div_total * rules["div_tax_foreign"]
            tax_note    = f"QSt ~{rules['div_tax_foreign']*100:.0f} % (DBA abhängig)"

        total_dividends    += div_total
        total_vst          += vst
        total_foreign_tax  += foreign_tax
        total_market_value += market_val

        rows.append({
            "ticker":      ticker,
            "name":        name[:28],
            "qty":         qty,
            "price":       price,
            "market_val":  market_val,
            "div_total":   div_total,
            "vst":         vst,
            "foreign_tax": foreign_tax,
            "tax_note":    tax_note,
        })

    summary = {
        "total_market_value": total_market_value,
        "total_dividends":    total_dividends,
        "total_vst":          total_vst,
        "total_foreign_tax":  total_foreign_tax,
        "total_tax":          total_vst + total_foreign_tax,
        "net_dividends":      total_dividends - total_vst - total_foreign_tax,
        "currency":           currency,
        "rules":              rules,
    }
    return rows, summary


def calc_de(positions, currency="EUR"):
    """DE: Abgeltungssteuer 25 % + Soli. Dividenden aus yfinance."""
    rows = []
    total_dividends = 0.0
    total_market_value = 0.0
    rules = TAX_RULES["DE"]
    frei  = rules["freistellungsauftrag"]
    year  = rules["year"]

    for pos in positions:
        ticker = pos.get("ticker", "")
        name   = pos.get("name", ticker)
        qty    = float(pos.get("quantity", 0))
        price  = float(pos.get("current_price", 0))
        buy_p  = float(pos.get("buy_price", price))

        market_val  = qty * price
        unrealised  = (price - buy_p) * qty

        if "_div_total" in pos:
            div_total = pos["_div_total"]
        else:
            raw = pos.get("raw_positions") or [{"quantity": qty, "buy_date": ""}]
            div_total, _ = _fetch_real_dividends(ticker, year, positions_for_sym=raw)

        total_dividends    += div_total
        total_market_value += market_val

        rows.append({
            "ticker":    ticker,
            "name":      name[:28],
            "qty":       qty,
            "price":     price,
            "market_val": market_val,
            "div_total": div_total,
            "unrealised": unrealised,
        })

    taxable = max(0.0, total_dividends - frei)
    abgeltungssteuer = taxable * rules["withholding_rate"]
    soli             = abgeltungssteuer * rules["soli_rate"]
    total_tax        = abgeltungssteuer + soli

    summary = {
        "total_market_value":  total_market_value,
        "total_dividends":     total_dividends,
        "freistellungsauftrag": frei,
        "taxable_dividends":   taxable,
        "abgeltungssteuer":    abgeltungssteuer,
        "soli":                soli,
        "total_tax":           total_tax,
        "net_dividends":       total_dividends - total_tax,
        "currency":            currency,
        "rules":               rules,
    }
    return rows, summary


def calc_uk(positions, currency="GBP"):
    """UK: Dividend Allowance £500, Basic Rate 8.75 %. Dividenden aus yfinance."""
    rows = []
    total_dividends = 0.0
    total_market_value = 0.0
    rules = TAX_RULES["UK"]
    year  = rules["year"]

    for pos in positions:
        ticker = pos.get("ticker", "")
        name   = pos.get("name", ticker)
        qty    = float(pos.get("quantity", 0))
        price  = float(pos.get("current_price", 0))
        buy_p  = float(pos.get("buy_price", price))

        market_val = qty * price
        unrealised = (price - buy_p) * qty

        if "_div_total" in pos:
            div_total = pos["_div_total"]
        else:
            raw = pos.get("raw_positions") or [{"quantity": qty, "buy_date": ""}]
            div_total, _ = _fetch_real_dividends(ticker, year, positions_for_sym=raw)

        total_dividends    += div_total
        total_market_value += market_val

        rows.append({
            "ticker":    ticker,
            "name":      name[:28],
            "qty":       qty,
            "price":     price,
            "market_val": market_val,
            "div_total": div_total,
            "unrealised": unrealised,
        })

    allowance = rules["dividend_allowance"]
    taxable   = max(0.0, total_dividends - allowance)
    div_tax   = taxable * rules["div_basic_rate"]

    summary = {
        "total_market_value": total_market_value,
        "total_dividends":    total_dividends,
        "dividend_allowance": allowance,
        "taxable_dividends":  taxable,
        "div_tax_basic":      div_tax,
        "total_tax":          div_tax,
        "net_dividends":      total_dividends - div_tax,
        "currency":           currency,
        "rules":              rules,
    }
    return rows, summary


def calc_us(positions, currency="USD"):
    """US: QSt 30 % / DBA CH-US 15 %. Dividenden aus yfinance."""
    rows = []
    total_dividends = 0.0
    total_tax       = 0.0
    total_market_value = 0.0
    rules = TAX_RULES["US"]
    year  = rules["year"]

    for pos in positions:
        ticker = pos.get("ticker", "")
        name   = pos.get("name", ticker)
        qty    = float(pos.get("quantity", 0))
        price  = float(pos.get("current_price", 0))
        country_of_investor = pos.get("investor_country", "CH").upper()

        market_val = qty * price

        if "_div_total" in pos:
            div_total = pos["_div_total"]
        else:
            raw = pos.get("raw_positions") or [{"quantity": qty, "buy_date": ""}]
            div_total, _ = _fetch_real_dividends(ticker, year, positions_for_sym=raw)

        if country_of_investor == "CH":
            rate = rules["dba_ch_rate"]
            rate_note = "DBA CH-US 15 %"
        elif country_of_investor == "DE":
            rate = rules["dba_de_rate"]
            rate_note = "DBA DE-US 15 %"
        else:
            rate = rules["nra_withholding"]
            rate_note = "Standard 30 %"

        tax = div_total * rate

        total_dividends    += div_total
        total_tax          += tax
        total_market_value += market_val

        rows.append({
            "ticker":    ticker,
            "name":      name[:28],
            "qty":       qty,
            "price":     price,
            "market_val": market_val,
            "div_total": div_total,
            "tax":       tax,
            "rate_note": rate_note,
        })

    summary = {
        "total_market_value": total_market_value,
        "total_dividends":    total_dividends,
        "total_tax":          total_tax,
        "net_dividends":      total_dividends - total_tax,
        "currency":           currency,
        "rules":              rules,
        "wash_sale_note": (
            "⚠️  Wash Sale Rule (IRC § 1091): Verluste aus Verkäufen, bei denen dieselbe "
            "Aktie innerhalb von 30 Tagen vor/nach dem Verkauf zurückgekauft wurde, "
            "sind steuerlich NICHT abzugsfähig. Bitte prüfen Sie Ihren Broker-Auszug "
            "(1099-B) auf 'Wash Sale Loss Disallowed'-Einträge."
        ),
    }
    return rows, summary



def calc_at(positions, currency="EUR"):
    """
    AT-Steuerauszug:
    - KESt 27,5 % flat auf Dividenden (automatisch durch Depotbank)
    - KESt 27,5 % auf realisierte Kursgewinne
    - DBA-Anrechnung für ausländische Quellensteuer
    Dividenden: tatsächlich bezahlte Beträge aus yfinance (Steuerjahr lt. TAX_RULES)
    """
    rows = []
    total_dividends    = 0.0
    total_kest_div     = 0.0
    total_kest_gain    = 0.0
    total_foreign_tax  = 0.0
    total_market_value = 0.0
    rules = TAX_RULES["AT"]
    year  = rules["year"]
    rate  = rules["kest_rate"]

    for pos in positions:
        ticker  = pos.get("ticker", "")
        name    = pos.get("name", ticker)
        qty     = float(pos.get("quantity", 0))
        price   = float(pos.get("current_price", 0))
        buy_p   = float(pos.get("buy_price", price))
        country = pos.get("country", "").upper()

        market_val  = qty * price
        unrealised  = (price - buy_p) * qty
        kest_gain   = max(0.0, unrealised) * rate  # nur auf realisierten Gewinn fällig

        if "_div_total" in pos:
            div_total = pos["_div_total"]
        else:
            raw = pos.get("raw_positions") or [{"quantity": qty, "buy_date": ""}]
            div_total, _ = _fetch_real_dividends(ticker, year, positions_for_sym=raw)

        # Quellensteuer-Anrechnung je Herkunftsland (DBA Österreich)
        if country in ("AT", "") or ticker.upper().endswith(".VI"):
            # Österreichische Wertpapiere: KESt wird direkt abgeführt, keine ausländ. QSt
            foreign_tax = 0.0
            kest_div    = div_total * rate
            tax_note    = "KESt 27,5 % (AT-Depot)"
        elif country == "US":
            # DBA AT-US: 15 % QSt anrechenbar auf KESt 27,5 %
            foreign_tax = div_total * 0.15
            kest_div    = max(0.0, div_total * rate - foreign_tax)
            tax_note    = "US QSt 15 % (DBA AT-US), Rest KESt"
        elif country == "DE":
            # DBA AT-DE: 15 % QSt
            foreign_tax = div_total * 0.15
            kest_div    = max(0.0, div_total * rate - foreign_tax)
            tax_note    = "DE QSt 15 % (DBA AT-DE), Rest KESt"
        elif country == "CH":
            # DBA AT-CH: 15 % QSt
            foreign_tax = div_total * 0.15
            kest_div    = max(0.0, div_total * rate - foreign_tax)
            tax_note    = "CH VSt 15 % (DBA AT-CH), Rest KESt"
        elif country == "GB":
            # DBA AT-UK: 15 % QSt
            foreign_tax = div_total * 0.15
            kest_div    = max(0.0, div_total * rate - foreign_tax)
            tax_note    = "UK QSt 15 % (DBA AT-UK), Rest KESt"
        elif country == "FR":
            # DBA AT-FR: 15 % QSt
            foreign_tax = div_total * 0.15
            kest_div    = max(0.0, div_total * rate - foreign_tax)
            tax_note    = "FR QSt 15 % (DBA AT-FR), Rest KESt"
        else:
            foreign_tax = div_total * 0.15   # Standardannahme DBA
            kest_div    = max(0.0, div_total * rate - foreign_tax)
            tax_note    = f"QSt ~15 % (DBA), Rest KESt {rate*100:.1f} %"

        total_dividends    += div_total
        total_kest_div     += kest_div
        total_kest_gain    += kest_gain
        total_foreign_tax  += foreign_tax
        total_market_value += market_val

        rows.append({
            "ticker":      ticker,
            "name":        name[:28],
            "qty":         qty,
            "price":       price,
            "market_val":  market_val,
            "div_total":   div_total,
            "kest_div":    kest_div,
            "foreign_tax": foreign_tax,
            "unrealised":  unrealised,
            "kest_gain":   kest_gain,
            "tax_note":    tax_note,
        })

    total_tax = total_kest_div + total_kest_gain + total_foreign_tax
    summary = {
        "total_market_value": total_market_value,
        "total_dividends":    total_dividends,
        "total_kest_div":     total_kest_div,
        "total_kest_gain":    total_kest_gain,
        "total_foreign_tax":  total_foreign_tax,
        "total_tax":          total_tax,
        "net_dividends":      total_dividends - total_kest_div - total_foreign_tax,
        "currency":           currency,
        "rules":              rules,
    }
    return rows, summary

CALC_FUNCS = {
    "AT": calc_at,
    "CH": calc_ch,
    "DE": calc_de,
    "UK": calc_uk,
    "US": calc_us,
}


# ── PyQt6 Dialog-Klasse ──────────────────────────────────────────────────────

def _build_tax_dialog(country_code, portfolio_data, parent=None, number_format="CH"):
    """
    Erstellt und gibt einen TaxDialog zurück.
    Wird von PortfolioDialog via show_tax_dialog() aufgerufen.
    """
    global _TAX_NUMBER_FORMAT
    _TAX_NUMBER_FORMAT = number_format
    from PyQt6.QtWidgets import (
        QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
        QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
        QFileDialog, QWidget, QSizePolicy, QScrollArea, QFrame,
    )
    from PyQt6.QtCore import Qt
    from PyQt6.QtGui  import QFont, QColor, QFontDatabase
    import os

    rules   = TAX_RULES[country_code]
    calc_fn = CALC_FUNCS[country_code]

    # ── Portfolio-Positionen extrahieren ─────────────────────────────────────
    # portfolio_data: {sym: {"quantity":..., "current_price":..., "buy_date":...,
    #                        "raw_positions": [{"quantity":q,"buy_date":d,...}]}}
    positions = []
    for sym, info in (portfolio_data or {}).items():
        qty   = info.get("quantity", 0)
        price = info.get("current_price") or info.get("buy_price", 0)
        buy   = info.get("buy_price", price)
        cty   = info.get("country", "")

        if qty and qty > 0:
            positions.append({
                "ticker":           sym,
                "name":             info.get("name", sym),
                "quantity":         qty,
                "current_price":    price,
                "buy_price":        buy,
                "country":          cty,
                "investor_country": country_code,
                # Rohdaten mit buy_date für Ex-Day-Berechnung
                "raw_positions":    info.get("raw_positions", []),
            })

    # ── Ladescreen: Dollarnoten-Animation + Status pro Symbol ─────────────────
    from PyQt6.QtWidgets import (QDialog as _QDlg, QVBoxLayout as _QVL,
                                  QLabel as _QLbl, QApplication as _QApp)
    from PyQt6.QtCore import Qt as _Qt2, QTimer as _QTimer
    from PyQt6.QtGui import QFont as _QFont

    load_dlg = _QDlg(parent)
    load_dlg.setWindowTitle(TRT("tax_loading_title"))
    load_dlg.setWindowFlags(
        _Qt2.WindowType.Dialog |
        _Qt2.WindowType.CustomizeWindowHint |
        _Qt2.WindowType.WindowTitleHint
    )
    load_dlg.setFixedSize(420, 200)
    load_dlg.setWindowModality(_Qt2.WindowModality.ApplicationModal)
    _lv = _QVL(load_dlg)
    _lv.setContentsMargins(20, 20, 20, 20)
    _lv.setSpacing(10)

    _emoji_lbl = _QLbl("💵")
    _emoji_lbl.setAlignment(_Qt2.AlignmentFlag.AlignCenter)
    _emoji_lbl.setFont(_QFont("Segoe UI Emoji", 40))
    _lv.addWidget(_emoji_lbl)

    _title_lbl = _QLbl(TRT("tax_loading_header"))
    _title_lbl.setAlignment(_Qt2.AlignmentFlag.AlignCenter)
    _lv.addWidget(_title_lbl)

    _status_lbl = _QLbl(TRT("tax_loading_status", i=0, n=len(positions)))
    _status_lbl.setAlignment(_Qt2.AlignmentFlag.AlignCenter)
    _status_lbl.setStyleSheet("color:#555; font-size:11px;")
    _lv.addWidget(_status_lbl)

    _sym_lbl = _QLbl("")
    _sym_lbl.setAlignment(_Qt2.AlignmentFlag.AlignCenter)
    _sym_lbl.setStyleSheet("color:#2980b9; font-size:11px; font-style:italic;")
    _lv.addWidget(_sym_lbl)

    # Dollarnoten rotieren: 💵 💴 💶 💷
    _notes = ["💵", "💴", "💶", "💷"]
    _note_idx = [0]
    def _rotate_note():
        _note_idx[0] = (_note_idx[0] + 1) % len(_notes)
        _emoji_lbl.setText(_notes[_note_idx[0]])
    _anim_timer = _QTimer()
    _anim_timer.setInterval(400)
    _anim_timer.timeout.connect(_rotate_note)
    _anim_timer.start()

    load_dlg.show()
    _QApp.processEvents()

    # Dividenden pro Symbol laden mit Live-Status
    def _calc_with_progress(positions):
        """Ruft calc_fn auf, aber aktualisiert den Ladescreen pro Symbol."""
        n = len(positions)
        for i, pos in enumerate(positions):
            sym = pos.get("ticker", "?")
            _status_lbl.setText(TRT("tax_loading_status", i=i, n=n))
            _sym_lbl.setText(TRT("tax_loading_symbol", sym=sym))
            _QApp.processEvents()
            # Dividenden für dieses Symbol holen
            raw = pos.get("raw_positions") or [{"quantity": pos.get("quantity",0), "buy_date":""}]
            year = rules["year"]
            pos["_div_total"], pos["_div_detail"] = _fetch_real_dividends(
                sym, year, positions_for_sym=raw)
        _status_lbl.setText(TRT("tax_loading_calculating", n=n))
        _sym_lbl.setText("")
        _QApp.processEvents()
        return calc_fn(positions)

    rows, summary = _calc_with_progress(positions)

    _anim_timer.stop()
    load_dlg.close()

    # ── Dialog aufbauen ───────────────────────────────────────────────────────
    dlg = QDialog(parent)
    dlg.setWindowTitle(TRT("tax_main_title", flag=rules['flag'], name=rules['name'], year=rules['year']))
    dlg.resize(1050, 680)
    dlg.setMinimumSize(820, 500)

    main_layout = QVBoxLayout(dlg)
    main_layout.setContentsMargins(12, 10, 12, 10)
    main_layout.setSpacing(8)

    # ── Kopfzeile: Update-Button links, dann Stretch, dann alle Aktions-Buttons ──
    header_row = QHBoxLayout()
    header_row.setSpacing(6)

    # Emoji-Font ermitteln (Windows: Segoe UI Emoji, Linux: Noto Color Emoji, Mac: Apple Color Emoji)
    _ef = None
    try:
        _af = QFontDatabase.families()
        for _fc in ("Segoe UI Emoji", "Noto Color Emoji", "Apple Color Emoji"):
            if _fc in _af:
                from PyQt6.QtGui import QFont as _QFontBtn
                _ef = _QFontBtn(_fc, 10)
                break
    except Exception:
        pass

    # Update-Button (Platzhalter – Funktion folgt nach Homepage-Launch)
    update_btn = QPushButton(TRT("tax_btn_update"))
    update_btn.setEnabled(False)   # Platzhalter – noch nicht aktiv
    update_btn.setStyleSheet("padding:4px 10px; color: gray;")
    if _ef:
        update_btn.setFont(_ef)
    header_row.addWidget(update_btn)

    header_row.addStretch()

    # Alle Aktions-Buttons oben rechts – gleiche Höhe und Style
    _BTN_H = 32  # einheitliche Button-Höhe (wie Export-Button)
    _BTN_STYLE = "padding:4px 14px;"

    save_btn   = QPushButton(TRT("tax_btn_save"))
    load_btn   = QPushButton(TRT("tax_btn_load"))
    export_btn = QPushButton(TRT("tax_btn_export"))
    export_btn.setToolTip(TRT("tax_tip_export"))
    export_btn.setStyleSheet(_BTN_STYLE + " font-weight:bold;")
    form_btn   = QPushButton(TRT("tax_btn_form"))
    form_btn.setToolTip(TRT("tax_tip_form_main"))
    form_btn.setStyleSheet(
        _BTN_STYLE + " background:#1a6b3a; color:white; font-weight:bold;"
    )
    if country_code not in ("CH", "DE", "AT", "UK", "US"):
        form_btn.setEnabled(False)
        form_btn.setToolTip(TRT("tax_tip_form_disabled"))
        form_btn.setStyleSheet(_BTN_STYLE + " background:#aaa; color:white;")
    close_btn  = QPushButton(TRT("tax_btn_close"))
    close_btn.setDefault(True)

    help_btn = QPushButton(TRT("tax_btn_help"))
    help_btn.setToolTip(TRT("tax_tip_help"))

    for b in (save_btn, load_btn, export_btn, form_btn, help_btn, close_btn):
        b.setMinimumHeight(_BTN_H)
        if b not in (export_btn, form_btn):
            b.setStyleSheet(_BTN_STYLE)
        if _ef:
            b.setFont(_ef)
        header_row.addWidget(b)

    main_layout.addLayout(header_row)

    # Titel
    title_lbl = QLabel(
        TRT("tax_main_heading", flag=rules['flag'], name=rules['name'], year=rules['year'])
    )
    title_lbl.setStyleSheet("font-size:15px; margin-bottom:2px;")
    main_layout.addWidget(title_lbl)

    # Disclaimer (Regeln)
    _disc_key = "disclaimer_en" if get_tax_language() == "EN" and "disclaimer_en" in rules else "disclaimer"
    rule_lbl = QLabel(rules[_disc_key])
    rule_lbl.setWordWrap(True)
    rule_lbl.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
    rule_lbl.setStyleSheet(
        "background:#f0f4ff; border:1px solid #c0c8e8; border-radius:4px;"
        "padding:6px 8px; font-size:11px; color:#334;"
    )
    main_layout.addWidget(rule_lbl)

    # ── Positions-Tabelle ─────────────────────────────────────────────────────
    _c = TRT
    if country_code == "CH":
        headers = [_c("tax_col_ticker"), _c("tax_col_name"), _c("tax_col_qty"),
                   _c("tax_col_price"), _c("tax_col_market_val"),
                   _c("tax_col_div"), _c("tax_col_vst"), _c("tax_col_foreign_tax"),
                   _c("tax_col_note")]
    elif country_code in ("DE", "UK"):
        headers = [_c("tax_col_ticker"), _c("tax_col_name"), _c("tax_col_qty"),
                   _c("tax_col_price"), _c("tax_col_market_val"),
                   _c("tax_col_div"), _c("tax_col_unrealised")]
    else:  # AT / US
        headers = [_c("tax_col_ticker"), _c("tax_col_name"), _c("tax_col_qty"),
                   _c("tax_col_price"), _c("tax_col_market_val"),
                   _c("tax_col_div"), _c("tax_col_withholding"), _c("tax_col_dba")]

    tbl = QTableWidget(len(rows), len(headers))
    tbl.setHorizontalHeaderLabels(headers)
    tbl.setAlternatingRowColors(True)
    tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
    tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
    tbl.horizontalHeader().setStretchLastSection(True)
    tbl.verticalHeader().setVisible(False)
    tbl.setMinimumHeight(220)

    cur = summary.get("currency", "–")

    def _num(val, decimals=2, prefix=""):
        try:
            return f"{prefix}{_fmt(float(val), decimals)}"
        except Exception:
            return "–"

    for r, row in enumerate(rows):
        def _item(text, align=Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter):
            it = QTableWidgetItem(str(text))
            it.setTextAlignment(align)
            return it

        left = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        tbl.setItem(r, 0, _item(row["ticker"],      left))
        tbl.setItem(r, 1, _item(row["name"],         left))
        tbl.setItem(r, 2, _item(_num(row["qty"], 0)))
        tbl.setItem(r, 3, _item(_num(row["price"])))
        tbl.setItem(r, 4, _item(_num(row["market_val"])))
        tbl.setItem(r, 5, _item(_num(row["div_total"])))

        if country_code == "AT":
            tbl.setItem(r, 6, _item(_num(row["kest_div"])))
            tbl.setItem(r, 7, _item(_num(row["foreign_tax"])))
            tbl.setItem(r, 8, _item(row["tax_note"], left))
        elif country_code == "CH":
            tbl.setItem(r, 6, _item(_num(row["vst"])))
            tbl.setItem(r, 7, _item(_num(row["foreign_tax"])))
            tbl.setItem(r, 8, _item(row["tax_note"], left))
        elif country_code in ("DE", "UK"):
            tbl.setItem(r, 6, _item(_num(row.get("unrealised", 0))))
        elif country_code == "US":
            tbl.setItem(r, 6, _item(_num(row.get("tax", 0))))
            tbl.setItem(r, 7, _item(row.get("rate_note", ""), left))

    tbl.resizeColumnsToContents()
    main_layout.addWidget(tbl)

    # ── Zusammenfassung ───────────────────────────────────────────────────────
    sumframe = QFrame()
    sumframe.setFrameShape(QFrame.Shape.StyledPanel)
    sumframe.setStyleSheet(
        "QFrame{background:#f8f9fa; border:1px solid #dde; border-radius:5px;}"
        "QLabel{font-size:12px;}"
    )
    sum_layout = QVBoxLayout(sumframe)
    sum_layout.setContentsMargins(12, 8, 12, 8)
    sum_layout.setSpacing(3)

    def _sum_row(label, value, bold=False, color=None):
        lbl = QLabel(f"<b>{label}</b>  {value}" if bold else f"{label}  {value}")
        if color:
            lbl.setStyleSheet(f"color:{color};")
        sum_layout.addWidget(lbl)

    _sum_row(TRT("tax_sum_market_val"),
             _num(summary["total_market_value"]) + f"  {cur}", bold=True)
    _sum_row(TRT("tax_sum_gross_div"),
             _num(summary["total_dividends"]) + f"  {cur}")

    if country_code == "AT":
        _sum_row(TRT("tax_sum_kest_div"),
                 _num(summary["total_kest_div"]) + f"  {cur}", color="#c0392b")
        _sum_row(TRT("tax_sum_foreign_tax"),
                 _num(summary["total_foreign_tax"]) + f"  {cur}", color="#c0392b")
        _sum_row(TRT("tax_sum_kest_gain"),
                 _num(summary["total_kest_gain"]) + f"  {cur}", color="#e67e22")
        _sum_row(TRT("tax_sum_total_tax"),
                 _num(summary["total_tax"]) + f"  {cur}", bold=True, color="#c0392b")
        _sum_row(TRT("tax_sum_net_div_kest"),
                 _num(summary["net_dividends"]) + f"  {cur}", bold=True, color="#27ae60")
        if rules.get("verlustausgleich"):
            _sum_row(TRT("tax_sum_verlustausgleich"), rules["verlustausgleich_note"])
    elif country_code == "CH":
        _sum_row(TRT("tax_sum_vst"),
                 _num(summary["total_vst"]) + f"  {cur}", color="#c0392b")
        _sum_row(TRT("tax_sum_foreign_tax_est"),
                 _num(summary["total_foreign_tax"]) + f"  {cur}", color="#c0392b")
        _sum_row(TRT("tax_sum_total_tax_ch"),
                 _num(summary["total_tax"]) + f"  {cur}", bold=True, color="#c0392b")
        _sum_row(TRT("tax_sum_net_div_ch"),
                 _num(summary["net_dividends"]) + f"  {cur}", bold=True, color="#27ae60")
        if rules.get("wealth_tax"):
            _sum_row(TRT("tax_sum_wealth_tax"), rules["wealth_tax_note"])
    elif country_code == "DE":
        _sum_row(TRT("tax_sum_freistellung"),
                 _num(summary["freistellungsauftrag"]) + f"  {cur}")
        _sum_row(TRT("tax_sum_taxable_div"),
                 _num(summary["taxable_dividends"]) + f"  {cur}")
        _sum_row(TRT("tax_sum_abgeltung"),
                 _num(summary["abgeltungssteuer"]) + f"  {cur}", color="#c0392b")
        _sum_row(TRT("tax_sum_soli"),
                 _num(summary["soli"]) + f"  {cur}", color="#c0392b")
        _sum_row(TRT("tax_sum_total_tax_de"),
                 _num(summary["total_tax"]) + f"  {cur}", bold=True, color="#c0392b")
        _sum_row(TRT("tax_sum_net_div_de"),
                 _num(summary["net_dividends"]) + f"  {cur}", bold=True, color="#27ae60")
    elif country_code == "UK":
        _sum_row(TRT("tax_sum_div_allowance"),
                 _num(summary["dividend_allowance"]) + "  GBP")
        _sum_row(TRT("tax_sum_taxable_div_uk"),
                 _num(summary["taxable_dividends"]) + "  GBP")
        _sum_row(TRT("tax_sum_div_tax_basic"),
                 _num(summary["div_tax_basic"]) + "  GBP", color="#c0392b")
        _sum_row(TRT("tax_sum_net_div_uk"),
                 _num(summary["net_dividends"]) + "  GBP", bold=True, color="#27ae60")
    elif country_code == "US":
        _sum_row(TRT("tax_sum_total_tax_us"),
                 _num(summary["total_tax"]) + f"  {cur}", bold=True, color="#c0392b")
        _sum_row(TRT("tax_sum_net_div_us"),
                 _num(summary["net_dividends"]) + f"  {cur}", bold=True, color="#27ae60")

    main_layout.addWidget(sumframe)

    # ── Wash Sale Hinweis (nur US) ────────────────────────────────────────────
    if country_code == "US":
        _ws_frame = QWidget()
        _ws_frame.setStyleSheet(
            "QWidget { background:#fff3cd; border:1px solid #e6a817; border-radius:4px; }"
        )
        _ws_hl = QHBoxLayout(_ws_frame)
        _ws_hl.setContentsMargins(10, 6, 10, 6)
        _ws_hl.setSpacing(8)
        _ws_icon = QLabel("⚠️")
        _ws_icon.setStyleSheet("border:none; background:transparent; font-size:13px;")
        _ws_icon.setAlignment(Qt.AlignmentFlag.AlignTop)
        if _ef:
            _ws_icon.setFont(_ef)
        _ws_hl.addWidget(_ws_icon, 0, Qt.AlignmentFlag.AlignTop)
        _ws_lbl = QLabel(TRT("tax_wash_sale_text"))
        _ws_lbl.setWordWrap(True)
        _ws_lbl.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        _ws_lbl.setStyleSheet("border:none; background:transparent; font-size:10px; color:#7d4800;")
        _ws_hl.addWidget(_ws_lbl, 1)
        main_layout.addWidget(_ws_frame)

    # ── Globaler Disclaimer ────────────────────────────────────────────────────
    _disc_frame = QWidget()
    _disc_frame.setStyleSheet(
        "QWidget { background:#fff8e1; border:1px solid #f0c040; border-radius:4px; }"
    )
    _disc_hl = QHBoxLayout(_disc_frame)
    _disc_hl.setContentsMargins(10, 6, 10, 6)
    _disc_hl.setSpacing(8)
    _disc_icon = QLabel("⚠️")
    _disc_icon.setStyleSheet("border:none; background:transparent; font-size:13px;")
    _disc_icon.setAlignment(Qt.AlignmentFlag.AlignTop)
    if _ef:
        _disc_icon.setFont(_ef)
    _disc_icon.setSizePolicy(_disc_icon.sizePolicy().horizontalPolicy(),
                              _disc_icon.sizePolicy().verticalPolicy())
    _disc_hl.addWidget(_disc_icon, 0, Qt.AlignmentFlag.AlignTop)
    disc_lbl = QLabel(TRT("tax_disclaimer_global"))
    disc_lbl.setWordWrap(True)
    disc_lbl.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
    disc_lbl.setStyleSheet("border:none; background:transparent; font-size:10px; color:#555;")
    _disc_hl.addWidget(disc_lbl, 1)
    main_layout.addWidget(_disc_frame)

    # Buttons sind in der Kopfzeile (oben) – kein separater btn_row mehr nötig

    # ── Interner Daten-State (für Export / Speichern) ──────────────────────────
    _tax_data = {
        "country":   country_code,
        "rules":     rules,
        "rows":      rows,
        "summary":   summary,
        "positions": positions,
    }

    # ── Export ────────────────────────────────────────────────────────────────
    def _on_export():
        from PyQt6.QtWidgets import QDialog as _QD, QVBoxLayout as _VL, QHBoxLayout as _HL
        from PyQt6.QtWidgets import QRadioButton, QButtonGroup, QLabel as _QL
        from PyQt6.QtWidgets import QPushButton as _QB

        fmt_dlg = _QD(dlg)
        fmt_dlg.setWindowTitle(TRT("tax_export_title"))
        fmt_dlg.setFixedWidth(320)
        lay = _VL(fmt_dlg)
        lay.setSpacing(10); lay.setContentsMargins(16, 14, 16, 14)

        lay.addWidget(_QL(TRT("tax_export_heading", name=rules['name'])))
        lay.addWidget(_QL(TRT("tax_export_format_label")))
        grp     = QButtonGroup(fmt_dlg)
        rb_pdf  = QRadioButton(TRT("tax_export_rb_pdf"))
        rb_xlsx = QRadioButton(TRT("tax_export_rb_xlsx"))
        rb_ods  = QRadioButton(TRT("tax_export_rb_ods"))
        rb_pdf.setChecked(True)
        for rb in (rb_pdf, rb_xlsx, rb_ods):
            grp.addButton(rb); lay.addWidget(rb)

        br = _HL()
        ok  = _QB(TRT("tax_export_btn_save")); ok.setStyleSheet("font-weight:bold;")
        can = _QB(TRT("tax_export_btn_cancel"))
        br.addWidget(ok); br.addWidget(can)
        lay.addLayout(br)
        can.clicked.connect(fmt_dlg.reject)

        def do_save():
            if rb_pdf.isChecked():    ext, flt = ".pdf",  "PDF (*.pdf)"
            elif rb_xlsx.isChecked(): ext, flt = ".xlsx", "Excel (*.xlsx)"
            else:                     ext, flt = ".ods",  "ODS (*.ods)"

            _default_name = TRT("tax_export_default_filename", country=country_code)
            path, _ = QFileDialog.getSaveFileName(
                dlg, TRT("tax_export_save_dialog"),
                os.path.expanduser(f"~/{_default_name}{ext}"), flt)
            if not path: return
            fmt_dlg.accept()
            try:
                if ext == ".pdf":
                    _export_pdf(path, _tax_data)
                elif ext == ".xlsx":
                    _export_xlsx(path, _tax_data)
                else:
                    _export_ods(path, _tax_data)
                QMessageBox.information(dlg, TRT("tax_export_success_title"),
                    TRT("tax_export_success_msg", path=path))
            except Exception as e:
                QMessageBox.warning(dlg, TRT("tax_export_error_title"),
                    TRT("tax_export_error_msg", error=e))

        ok.clicked.connect(do_save)
        import sys as _sys_fmt
        if _sys_fmt.platform == 'win32':
            fmt_dlg.finished.connect(lambda: (dlg.raise_(), dlg.activateWindow()))
        fmt_dlg.exec()

    export_btn.clicked.connect(_on_export)

    # ── Steuerformular-Export ─────────────────────────────────────────────────
    def _on_form_export():
        if not _tax_data.get("rows"):
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(dlg, TRT("tax_no_data_title"),
                TRT("tax_no_data_msg"))
            return
        _export_tax_form(dlg, _tax_data, country_code)

    form_btn.clicked.connect(_on_form_export)

    # ── Speichern (verschlüsselt) ─────────────────────────────────────────────
    def _on_save():
        _save_tax_data(dlg, _tax_data, country_code)

    save_btn.clicked.connect(_on_save)

    # ── Laden ─────────────────────────────────────────────────────────────────
    def _on_load():
        _load_tax_data(dlg, _tax_data, tbl, sumframe)

    load_btn.clicked.connect(_on_load)
    close_btn.clicked.connect(dlg.close)

    # ── Hilfe-Button ──────────────────────────────────────────────────────────
    _CSS = """<style>
body { font-family: sans-serif; font-size: 13px; }
h3 { color: #2c5f8a; margin-top: 14px; margin-bottom: 4px; }
h4 { color: #444; margin-top: 10px; margin-bottom: 3px; }
table { border-collapse: collapse; width: 100%; margin: 6px 0; font-size: 12px; }
th { background: #dce8f5; padding: 4px 7px; text-align: left; }
td { padding: 3px 7px; border-bottom: 1px solid #e0e0e0; }
ul { margin: 4px 0 6px 0; padding-left: 18px; }
li { margin-bottom: 3px; }
.tip { background: #e8f5e9; border-left: 4px solid #43a047; padding: 6px 10px; margin: 8px 0; border-radius: 3px; }
.warn { background: #fff3e0; border-left: 4px solid #e65100; padding: 6px 10px; margin: 8px 0; border-radius: 3px; }
code { background: #f5f5f5; padding: 1px 4px; border-radius: 3px; font-size: 12px; }
</style>"""

    _HELP_TEXTS = {
        "CH": {
            "DE": _CSS + """
        <h3>&#127464;&#127469; Steuermodul Schweiz (CH)</h3>
        <p>Das Schweizer Steuermodul erstellt einen <b>steuerrelevanten Jahresauszug</b> für das aktive Portfolio
        und berechnet Dividenden ex-dividend-day-genau, Verrechnungssteuer sowie ausländische Quellensteuern nach DBA.</p>
        <h3>Übersicht</h3>
        <table>
            <tr><th>Währung</th><th>Steuersystem</th><th>Besonderheiten</th></tr>
            <tr><td>CHF</td><td>Einkommens- + Vermögenssteuer</td><td>Verrechnungssteuer 35 % (rückforderbar) &bull; DBA CH–US/DE/FR/GB &bull; Formular DA-1 &bull; Wertschriftenverzeichnis R-US</td></tr>
        </table>
        <h3>Spalten im Steuerauszug</h3>
        <table>
            <tr><th>Spalte</th><th>Erklärung</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Wertpapierbezeichnung</td></tr>
            <tr><td><b>Stück</b></td><td>Gehaltene Stückzahl am Ex-Dividend-Tag</td></tr>
            <tr><td><b>Kurs / Marktwert</b></td><td>Kurs per 31.12. in CHF (SNB-Jahreskurs)</td></tr>
            <tr><td><b>Dividende/J.</b></td><td>Brutto-Dividende im Steuerjahr in CHF</td></tr>
            <tr><td><b>Verrechnungssteuer 35 %</b></td><td>Einbehaltene VSt auf Schweizer Dividenden (rückforderbar)</td></tr>
            <tr><td><b>Ausländische QSt</b></td><td>Einbehaltene Quellensteuer auf ausländische Dividenden (anrechenbar via DA-1)</td></tr>
            <tr><td><b>DBA-Satz / Hinweis</b></td><td>Angewandter DBA-Satz je Herkunftsland</td></tr>
        </table>
        <h3>Zusammenfassung (unterhalb der Tabelle)</h3>
        <table>
            <tr><th>Kennzahl</th><th>Bedeutung</th></tr>
            <tr><td>Gesamter Marktwert</td><td>Depotgesamtwert per 31.12. in CHF → Vermögenssteuer-Basis</td></tr>
            <tr><td>Brutto-Dividenden</td><td>Total aller Dividenden vor Abzug der Steuern</td></tr>
            <tr><td>Verrechnungssteuer 35 %</td><td>Rückforderbar via Steuererklärung (Formular 25)</td></tr>
            <tr><td>Ausländische QSt (anrechenbar)</td><td>Auf direkte Bundessteuer anrechenbar via DA-1</td></tr>
            <tr><td>Netto-Dividenden</td><td>Dividenden nach Abzug aller Steuern</td></tr>
        </table>
        <h3>DBA und Verrechnungssteuer im Detail</h3>
        <ul>
            <li><b>Verrechnungssteuer (VSt) 35 %</b>: Gilt auf Schweizer Dividenden (z.B. NESN.SW, NOVN.SW). Vollständig <b>rückforderbar</b> via Formular 25.</li>
            <li><b>Ausländische QSt</b>: Für US-Aktien 15 % dank DBA CH–US. Anrechenbar via <b>Formular DA-1</b>.</li>
            <li>Der Übertrag in das <b>kantonale Wertschriftenverzeichnis</b> ist im Steuerformular-PDF Zeile für Zeile vorbereitet.</li>
        </ul>
        <h3>Steuerformular-Export: Wertschriftenverzeichnis (DA-1 / R-US)</h3>
        <ul>
            <li><b>Auswahl:</b> Kanton (alle 26)</li>
            <li>Marktwert per 31.12. in CHF &bull; Dividenden brutto &bull; VSt 35 % &bull; Ausländische QSt &bull; SNB-Jahreskurse</li>
        </ul>
        <div class="tip"><b>Tipp:</b> Das PDF zeigt zu jeder Zahl das exakte <b>Formularfeld</b> (z.B. «DA-1 Feld 4»).</div>
        <div class="tip"><b>Tipp:</b> <b>🔄 Update Steuerregeln</b> lädt die aktuellen Steuersätze neu.</div>
        <h3>Bekannte Einschränkungen</h3>
        <ul>
            <li>Dividenden aus Yahoo Finance – bei REITs, ETFs können Daten unvollständig sein</li>
            <li>Realisierte Kursgewinne werden nicht berechnet</li>
            <li>Thesaurierende ETFs werden nicht erfasst</li>
            <li>Abweichungen vom Banksteuerauszug ca. 1–3 % möglich</li>
        </ul>
        <div class="warn">&#9888; <b>Orientierungshilfe</b> – keine rechtsgültige Steuererklärung.
        Offizielle Abgabe via kantonales Steuerportal. Bei Unsicherheiten Steuerberater konsultieren.</div>""",

            "EN": _CSS + """
        <h3>&#127464;&#127469; Tax Module Switzerland (CH)</h3>
        <p>The Swiss tax module generates a <b>tax-relevant annual statement</b> for the active portfolio,
        calculating dividends ex-dividend-day-accurately, withholding tax (Verrechnungssteuer), and foreign withholding taxes under DTTs.</p>
        <h3>Overview</h3>
        <table>
            <tr><th>Currency</th><th>Tax system</th><th>Key features</th></tr>
            <tr><td>CHF</td><td>Income tax + wealth tax</td><td>Withholding tax 35% (reclaimable) &bull; DTT CH–US/DE/FR/GB &bull; Form DA-1 &bull; Securities register R-US</td></tr>
        </table>
        <h3>Columns in the Tax Statement</h3>
        <table>
            <tr><th>Column</th><th>Explanation</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Security name</td></tr>
            <tr><td><b>Shares</b></td><td>Shares held on ex-dividend date</td></tr>
            <tr><td><b>Price / Market value</b></td><td>Price as of 31 Dec in CHF (SNB annual rate)</td></tr>
            <tr><td><b>Dividend/Y.</b></td><td>Gross dividend in the tax year in CHF</td></tr>
            <tr><td><b>Withholding tax 35%</b></td><td>Swiss withholding tax on Swiss dividends (reclaimable)</td></tr>
            <tr><td><b>Foreign WHT</b></td><td>Foreign withholding tax on foreign dividends (creditable via DA-1)</td></tr>
            <tr><td><b>DTT rate / Note</b></td><td>Applied DTT rate by source country</td></tr>
        </table>
        <h3>Summary (below the table)</h3>
        <table>
            <tr><th>Metric</th><th>Meaning</th></tr>
            <tr><td>Total market value</td><td>Portfolio total as of 31 Dec in CHF → wealth tax base</td></tr>
            <tr><td>Gross dividends</td><td>Total dividends before any deductions</td></tr>
            <tr><td>Withholding tax 35%</td><td>Reclaimable via tax return (Form 25)</td></tr>
            <tr><td>Foreign WHT (creditable)</td><td>Creditable against federal direct tax via DA-1</td></tr>
            <tr><td>Net dividends</td><td>Dividends after all taxes</td></tr>
        </table>
        <h3>DTT and Withholding Tax in Detail</h3>
        <ul>
            <li><b>Withholding tax (VSt) 35%</b>: Applies to Swiss dividends (e.g. NESN.SW). Fully <b>reclaimable</b> via Form 25.</li>
            <li><b>Foreign WHT</b>: For US stocks 15% under DTT CH–US. Creditable via <b>Form DA-1</b>.</li>
            <li>The transfer to the <b>cantonal securities register</b> is prepared line by line in the tax form PDF.</li>
        </ul>
        <h3>Tax Form Export: Securities Register (DA-1 / R-US)</h3>
        <ul>
            <li><b>Selection:</b> Canton (all 26)</li>
            <li>Market value as of 31 Dec in CHF &bull; Gross dividends &bull; WHT 35% &bull; Foreign WHT &bull; SNB annual rates</li>
        </ul>
        <div class="tip"><b>Tip:</b> The PDF shows the exact <b>form field</b> for each value (e.g. "DA-1 Field 4").</div>
        <div class="tip"><b>Tip:</b> <b>🔄 Update tax rules</b> reloads current tax rates.</div>
        <h3>Known Limitations</h3>
        <ul>
            <li>Dividends from Yahoo Finance – data may be incomplete for REITs or ETFs</li>
            <li>Realised capital gains are not calculated</li>
            <li>Accumulating ETFs are not captured</li>
            <li>Deviations from official bank statement approx. 1–3% possible</li>
        </ul>
        <div class="warn">&#9888; <b>Guidance tool only</b> – not a legally binding tax return.
        File officially via your cantonal tax portal. Consult a tax advisor if in doubt.</div>""",
        },

        "DE": {
            "DE": _CSS + """
        <h3>&#127465;&#127466; Steuermodul Deutschland (DE)</h3>
        <p>Das deutsche Steuermodul erstellt einen <b>steuerrelevanten Jahresauszug</b> für das aktive Portfolio
        und berechnet Dividenden ex-dividend-day-genau, Abgeltungsteuer, Solidaritätszuschlag sowie ausländische Quellensteuern nach DBA.</p>
        <h3>Übersicht</h3>
        <table>
            <tr><th>Währung</th><th>Steuersystem</th><th>Besonderheiten</th></tr>
            <tr><td>EUR</td><td>Abgeltungsteuer</td><td>25 % + Soli 5,5 % &bull; Freistellungsauftrag 1'000 € p.a. &bull; Kirchensteuer je Bundesland &bull; Jahressteuerbescheinigung der Depotbank</td></tr>
        </table>
        <h3>Spalten im Steuerauszug</h3>
        <table>
            <tr><th>Spalte</th><th>Erklärung</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Wertpapierbezeichnung</td></tr>
            <tr><td><b>Stück</b></td><td>Gehaltene Stückzahl am Ex-Dividend-Tag</td></tr>
            <tr><td><b>Kurs / Marktwert</b></td><td>Kurs per 31.12. in EUR (Bundesbank-Jahreskurs)</td></tr>
            <tr><td><b>Dividende/J.</b></td><td>Brutto-Dividende im Steuerjahr in EUR</td></tr>
            <tr><td><b>Ausländische QSt</b></td><td>Einbehaltene Quellensteuer auf ausländische Dividenden (anrechenbar)</td></tr>
            <tr><td><b>Abgeltungsteuer</b></td><td>25 % auf steuerpflichtige Kapitalerträge (nach Freistellungsauftrag)</td></tr>
            <tr><td><b>Nicht realisierter Gewinn</b></td><td>Buchgewinn der gehaltenen Position (informativ)</td></tr>
            <tr><td><b>DBA-Satz / Hinweis</b></td><td>Angewandter DBA-Satz je Herkunftsland</td></tr>
        </table>
        <h3>Zusammenfassung (unterhalb der Tabelle)</h3>
        <table>
            <tr><th>Kennzahl</th><th>Bedeutung</th></tr>
            <tr><td>Gesamter Marktwert</td><td>Depotgesamtwert per 31.12. in EUR</td></tr>
            <tr><td>Kapitalerträge brutto</td><td>Total aller Dividenden vor Abzug</td></tr>
            <tr><td>Freistellungsauftrag-Verbrauch</td><td>Wie viel des Freibetrags ist verbraucht</td></tr>
            <tr><td>Abgeltungsteuer 25 %</td><td>Auf den Freibetrag übersteigenden Betrag</td></tr>
            <tr><td>Solidaritätszuschlag 5,5 %</td><td>Auf die Abgeltungsteuer</td></tr>
            <tr><td>Steuerbelastung gesamt</td><td>Abgeltungsteuer + Soli</td></tr>
            <tr><td>Bereits einbehaltene QSt</td><td>Anrechenbare ausländische Quellensteuern</td></tr>
        </table>
        <h3>Freistellungsauftrag und Abgeltungsteuer im Detail</h3>
        <ul>
            <li><b>Freistellungsauftrag</b>: 1'000 € p.a. (Einzelperson) bzw. 2'000 € (Ehepaare). Das Modul berechnet den Verbrauch.</li>
            <li><b>Abgeltungsteuer 25 %</b> + <b>Soli 5,5 %</b> auf den übersteigenden Betrag.</li>
            <li><b>Kirchensteuer</b>: 8 % (BY, BW) oder 9 % (alle anderen) – im PDF als Hinweis, da Status unbekannt.</li>
            <li>Bei deutschem Depot: Depotbank führt Steuer automatisch ab (<b>Jahressteuerbescheinigung</b>). Modul primär für <b>Kontrolle und ausländische Depots</b> (z.B. IBKR).</li>
        </ul>
        <h3>Steuerformular-Export: Anlage KAP</h3>
        <ul>
            <li><b>Auswahl:</b> Bundesland (alle 16)</li>
            <li>Kapitalerträge Zeile 7 &bull; Freistellungsauftrag &bull; Abgeltungsteuer &bull; Soli &bull; Bundesbank-Jahreskurse</li>
        </ul>
        <div class="tip"><b>Tipp:</b> Das PDF zeigt das exakte <b>Formularfeld</b> (z.B. «Anlage KAP Zeile 7») für ELSTER.</div>
        <div class="tip"><b>Tipp:</b> <b>🔄 Update Steuerregeln</b> lädt die aktuellen Steuersätze neu.</div>
        <h3>Bekannte Einschränkungen</h3>
        <ul>
            <li>Dividenden aus Yahoo Finance – bei REITs, ETFs können Daten unvollständig sein</li>
            <li>Realisierte Kursgewinne werden nicht berechnet</li>
            <li>Thesaurierende ETFs werden nicht erfasst</li>
            <li>Abweichungen vom Banksteuerauszug ca. 1–3 % möglich</li>
        </ul>
        <div class="warn">&#9888; <b>Orientierungshilfe</b> – keine rechtsgültige Steuererklärung.
        Offizielle Abgabe via <b>ELSTER (elster.de)</b>. Bei Unsicherheiten Steuerberater konsultieren.</div>""",

            "EN": _CSS + """
        <h3>&#127465;&#127466; Tax Module Germany (DE)</h3>
        <p>The German tax module generates a <b>tax-relevant annual statement</b> for the active portfolio,
        calculating dividends ex-dividend-day-accurately, withholding tax (Abgeltungsteuer), solidarity surcharge (Soli), and foreign withholding taxes under DTTs.</p>
        <h3>Overview</h3>
        <table>
            <tr><th>Currency</th><th>Tax system</th><th>Key features</th></tr>
            <tr><td>EUR</td><td>Flat-rate withholding tax</td><td>25% + Soli 5.5% &bull; Tax-free allowance €1,000 p.a. &bull; Church tax by federal state &bull; Annual tax certificate from broker</td></tr>
        </table>
        <h3>Columns in the Tax Statement</h3>
        <table>
            <tr><th>Column</th><th>Explanation</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Security name</td></tr>
            <tr><td><b>Shares</b></td><td>Shares held on ex-dividend date</td></tr>
            <tr><td><b>Price / Market value</b></td><td>Price as of 31 Dec in EUR (Bundesbank annual rate)</td></tr>
            <tr><td><b>Dividend/Y.</b></td><td>Gross dividend in the tax year in EUR</td></tr>
            <tr><td><b>Foreign WHT</b></td><td>Foreign withholding tax on foreign dividends (creditable)</td></tr>
            <tr><td><b>Withholding tax</b></td><td>25% on taxable capital income (after tax-free allowance)</td></tr>
            <tr><td><b>Unrealised gain</b></td><td>Book gain on held position (informational)</td></tr>
            <tr><td><b>DTT rate / Note</b></td><td>Applied DTT rate by source country</td></tr>
        </table>
        <h3>Summary (below the table)</h3>
        <table>
            <tr><th>Metric</th><th>Meaning</th></tr>
            <tr><td>Total market value</td><td>Portfolio total as of 31 Dec in EUR</td></tr>
            <tr><td>Gross capital income</td><td>Total dividends before deductions</td></tr>
            <tr><td>Tax-free allowance used</td><td>How much of the allowance is consumed</td></tr>
            <tr><td>Withholding tax 25%</td><td>On amount exceeding the allowance</td></tr>
            <tr><td>Solidarity surcharge 5.5%</td><td>On top of withholding tax</td></tr>
            <tr><td>Total tax burden</td><td>Withholding tax + Soli</td></tr>
            <tr><td>Foreign WHT already withheld</td><td>Creditable foreign withholding taxes</td></tr>
        </table>
        <h3>Tax-free Allowance and Withholding Tax in Detail</h3>
        <ul>
            <li><b>Tax-free allowance (Freistellungsauftrag)</b>: €1,000 p.a. (individual) or €2,000 (married). The module calculates consumption.</li>
            <li><b>Withholding tax 25%</b> + <b>Soli 5.5%</b> on the amount exceeding the allowance.</li>
            <li><b>Church tax</b>: 8% (Bavaria, Baden-Württemberg) or 9% (all others) – shown as a note in the PDF.</li>
            <li>With a German broker: taxes are deducted automatically (annual tax certificate). Module mainly for <b>verification and foreign brokers</b> (e.g. IBKR).</li>
        </ul>
        <h3>Tax Form Export: Anlage KAP</h3>
        <ul>
            <li><b>Selection:</b> Federal state (all 16)</li>
            <li>Capital income line 7 &bull; Allowance offset &bull; Withholding tax &bull; Soli &bull; Bundesbank annual rates</li>
        </ul>
        <div class="tip"><b>Tip:</b> The PDF shows the exact <b>form field</b> (e.g. "Anlage KAP Line 7") for ELSTER entry.</div>
        <div class="tip"><b>Tip:</b> <b>🔄 Update tax rules</b> reloads current tax rates.</div>
        <h3>Known Limitations</h3>
        <ul>
            <li>Dividends from Yahoo Finance – data may be incomplete for REITs or ETFs</li>
            <li>Realised capital gains are not calculated</li>
            <li>Accumulating ETFs are not captured</li>
            <li>Deviations from official bank statement approx. 1–3% possible</li>
        </ul>
        <div class="warn">&#9888; <b>Guidance tool only</b> – not a legally binding tax return.
        File officially via <b>ELSTER (elster.de)</b>. Consult a tax advisor if in doubt.</div>""",
        },

        "AT": {
            "DE": _CSS + """
        <h3>&#127462;&#127481; Steuermodul Österreich (AT)</h3>
        <p>Das österreichische Steuermodul erstellt einen <b>steuerrelevanten Jahresauszug</b> und
        berechnet Dividenden ex-dividend-day-genau sowie die Kapitalertragsteuer (KESt) nach DBA.</p>
        <h3>Übersicht</h3>
        <table>
            <tr><th>Währung</th><th>Steuersystem</th><th>Besonderheiten</th></tr>
            <tr><td>EUR</td><td>Kapitalertragsteuer (KESt)</td><td>27,5 % flat &bull; Depotbank führt ab &bull; DBA 15 % anrechenbar &bull; Verlustausgleich &bull; E1kv-Beilage</td></tr>
        </table>
        <h3>Spalten im Steuerauszug</h3>
        <table>
            <tr><th>Spalte</th><th>Erklärung</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Wertpapierbezeichnung</td></tr>
            <tr><td><b>Stück</b></td><td>Gehaltene Stückzahl am Ex-Dividend-Tag</td></tr>
            <tr><td><b>Kurs / Marktwert</b></td><td>Kurs per 31.12. in EUR (OeNB-Jahreskurs)</td></tr>
            <tr><td><b>Dividende/J.</b></td><td>Brutto-Dividende im Steuerjahr in EUR</td></tr>
            <tr><td><b>Ausländische QSt</b></td><td>Einbehaltene Quellensteuer auf ausländische Dividenden</td></tr>
            <tr><td><b>KeSt 27,5 %</b></td><td>Kapitalertragsteuer auf Dividenden</td></tr>
            <tr><td><b>DBA-Satz / Hinweis</b></td><td>Anwendbarer DBA-Satz (AT–US/DE/CH/GB/FR: 15 %)</td></tr>
        </table>
        <h3>Steuerformular-Export: E1kv-Nachweis</h3>
        <ul>
            <li><b>Auswahl:</b> Finanzamt-Region (alle 9)</li>
            <li>KZ 369 (Dividenden) &bull; KeSt-Berechnung &bull; QSt anrechenbar &bull; KZ 802 (Depotwert) &bull; OeNB-Jahreskurse</li>
        </ul>
        <div class="tip"><b>Tipp:</b> Das PDF zeigt das exakte <b>Kennzahlfeld</b> (z.B. «KZ 369»).</div>
        <div class="warn">&#9888; <b>Orientierungshilfe</b> – offizielle Abgabe via <b>FinanzOnline</b>. Bei Unsicherheiten Steuerberater konsultieren.</div>""",

            "EN": _CSS + """
        <h3>&#127462;&#127481; Tax Module Austria (AT)</h3>
        <p>The Austrian tax module generates a <b>tax-relevant annual statement</b>,
        calculating dividends ex-dividend-day-accurately and capital gains tax (KESt) under DTTs.</p>
        <h3>Overview</h3>
        <table>
            <tr><th>Currency</th><th>Tax system</th><th>Key features</th></tr>
            <tr><td>EUR</td><td>Capital gains tax (KESt)</td><td>27.5% flat &bull; Withheld by broker &bull; DTT 15% creditable &bull; Loss offset &bull; E1kv supplement</td></tr>
        </table>
        <h3>Columns in the Tax Statement</h3>
        <table>
            <tr><th>Column</th><th>Explanation</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Security name</td></tr>
            <tr><td><b>Shares</b></td><td>Shares held on ex-dividend date</td></tr>
            <tr><td><b>Price / Market value</b></td><td>Price as of 31 Dec in EUR (OeNB annual rate)</td></tr>
            <tr><td><b>Dividend/Y.</b></td><td>Gross dividend in the tax year in EUR</td></tr>
            <tr><td><b>Foreign WHT</b></td><td>Foreign withholding tax on foreign dividends</td></tr>
            <tr><td><b>KeSt 27.5%</b></td><td>Capital gains tax on dividends</td></tr>
            <tr><td><b>DTT rate / Note</b></td><td>Applicable DTT rate (AT–US/DE/CH/GB/FR: 15%)</td></tr>
        </table>
        <h3>Tax Form Export: E1kv Supplement</h3>
        <ul>
            <li><b>Selection:</b> Tax office region (all 9)</li>
            <li>KZ 369 (dividends) &bull; KeSt calculation &bull; WHT creditable &bull; KZ 802 (portfolio value) &bull; OeNB annual rates</li>
        </ul>
        <div class="tip"><b>Tip:</b> The PDF shows the exact <b>field code</b> (e.g. "KZ 369") for FinanzOnline entry.</div>
        <div class="warn">&#9888; <b>Guidance tool only</b> – file officially via <b>FinanzOnline</b>. Consult a tax advisor if in doubt.</div>""",
        },

        "UK": {
            "DE": _CSS + """
        <h3>&#127468;&#127463; Steuermodul Grossbritannien (UK)</h3>
        <p>Das UK-Steuermodul erstellt einen <b>steuerrelevanten Jahresauszug</b> und
        berechnet Dividenden ex-dividend-day-genau, Dividend Tax sowie Capital Gains Tax (CGT).</p>
        <h3>Übersicht</h3>
        <table>
            <tr><th>Währung</th><th>Steuersystem</th><th>Besonderheiten</th></tr>
            <tr><td>GBP</td><td>Income Tax / CGT</td><td>Dividend Allowance £500 p.a. &bull; Basic Rate 8,75 % &bull; Higher Rate 33,75 % &bull; CGT Allowance £3'000 &bull; Steuerjahr 6. April – 5. April</td></tr>
        </table>
        <h3>Spalten im Steuerauszug</h3>
        <table>
            <tr><th>Spalte</th><th>Erklärung</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Wertpapierbezeichnung</td></tr>
            <tr><td><b>Stück</b></td><td>Gehaltene Stückzahl am Ex-Dividend-Tag</td></tr>
            <tr><td><b>Kurs / Marktwert</b></td><td>Kurs per 5. April in GBP</td></tr>
            <tr><td><b>Dividende/J.</b></td><td>Brutto-Dividende im Steuerjahr in GBP</td></tr>
            <tr><td><b>Nicht realisierter Gewinn</b></td><td>Buchgewinn (informativ, CGT erst bei Verkauf)</td></tr>
        </table>
        <h3>Steuerformular-Export: SA100 / SA108 (Self Assessment)</h3>
        <ul>
            <li><b>Auswahl:</b> Nation (England/Wales, Schottland, Nordirland)</li>
            <li>Dividend Allowance &bull; Taxable dividends &bull; Div Tax &bull; CGT Exempt &bull; A4-Format</li>
        </ul>
        <div class="tip"><b>Tipp:</b> Das PDF zeigt das exakte <b>Formularfeld</b> für die HMRC Self Assessment Eingabe.</div>
        <div class="warn">&#9888; <b>Orientierungshilfe</b> – offizielle Abgabe via <b>HMRC Self Assessment (gov.uk)</b>. Bei Unsicherheiten Steuerberater konsultieren.</div>""",

            "EN": _CSS + """
        <h3>&#127468;&#127463; Tax Module United Kingdom (UK)</h3>
        <p>The UK tax module generates a <b>tax-relevant annual statement</b>,
        calculating dividends ex-dividend-day-accurately, Dividend Tax, and Capital Gains Tax (CGT).</p>
        <h3>Overview</h3>
        <table>
            <tr><th>Currency</th><th>Tax system</th><th>Key features</th></tr>
            <tr><td>GBP</td><td>Income Tax / CGT</td><td>Dividend Allowance £500 p.a. &bull; Basic Rate 8.75% &bull; Higher Rate 33.75% &bull; CGT Allowance £3,000 &bull; Tax year 6 April – 5 April</td></tr>
        </table>
        <h3>Columns in the Tax Statement</h3>
        <table>
            <tr><th>Column</th><th>Explanation</th></tr>
            <tr><td><b>Ticker / Name</b></td><td>Security name</td></tr>
            <tr><td><b>Shares</b></td><td>Shares held on ex-dividend date</td></tr>
            <tr><td><b>Price / Market value</b></td><td>Price as of 5 April in GBP</td></tr>
            <tr><td><b>Dividend/Y.</b></td><td>Gross dividend in the tax year in GBP</td></tr>
            <tr><td><b>Unrealised gain</b></td><td>Book gain (informational – CGT only on disposal)</td></tr>
        </table>
        <h3>Tax Form Export: SA100 / SA108 (Self Assessment)</h3>
        <ul>
            <li><b>Selection:</b> Nation (England/Wales, Scotland, Northern Ireland)</li>
            <li>Dividend Allowance &bull; Taxable dividends &bull; Div Tax &bull; CGT Exempt &bull; A4 format</li>
        </ul>
        <div class="tip"><b>Tip:</b> The PDF shows the exact <b>form field</b> for HMRC Self Assessment entry.</div>
        <div class="warn">&#9888; <b>Guidance tool only</b> – file officially via <b>HMRC Self Assessment (gov.uk)</b>. Consult a tax advisor if in doubt.</div>""",
        },

        "US": {
            "DE": _CSS + """
        <h3>&#127482;&#127480; Steuermodul USA (US)</h3>
        <p>Das US-Steuermodul berechnet Dividenden ex-dividend-day-genau, Quellensteuer (DBA oder 30 % Standard),
        Qualified Dividend Rates, State Tax und NIIT.</p>
        <h3>Übersicht</h3>
        <table>
            <tr><th>Währung</th><th>Steuersystem</th><th>Besonderheiten</th></tr>
            <tr><td>USD</td><td>Federal + State Tax</td><td>DBA CH/DE/AT–US 15 % &bull; Standard-QSt 30 % &bull; Qualified Dividends 0/15/20 % &bull; State Tax je Bundesstaat</td></tr>
        </table>
        <h3>Investor-Profile</h3>
        <table>
            <tr><th>Status</th><th>Formular</th><th>Quellensteuer</th></tr>
            <tr><td><b>CH / DE / AT</b></td><td>1042-S</td><td>15 % (DBA)</td></tr>
            <tr><td><b>NRA</b></td><td>1042-S</td><td>30 %</td></tr>
            <tr><td><b>US-Resident</b></td><td>Form 1040 / Schedule B</td><td>0/15/20 % je Filing Status</td></tr>
        </table>
        <h3>Filing Status und Einkommensschwellen 2024 (US-Resident)</h3>
        <table>
            <tr><th>Filing Status</th><th>0 %-Rate bis</th><th>15 %-Rate bis</th><th>20 % ab</th></tr>
            <tr><td>Single</td><td>$ 47'025</td><td>$ 518'900</td><td>&gt; $ 518'900</td></tr>
            <tr><td>Married Filing Jointly</td><td>$ 94'050</td><td>$ 583'750</td><td>&gt; $ 583'750</td></tr>
            <tr><td>Married Filing Separately</td><td>$ 47'025</td><td>$ 291'850</td><td>&gt; $ 291'850</td></tr>
            <tr><td>Head of Household</td><td>$ 63'000</td><td>$ 551'350</td><td>&gt; $ 551'350</td></tr>
        </table>
        <h3>Wash Sale Rule (IRC § 1091)</h3>
        <ul>
            <li>Gilt nur für <b>US-Steuerpflichtige</b>. Schweizer, Deutsche und Österreicher sind in der Regel <b>nicht betroffen</b>.</li>
            <li>Modul berechnet dies nicht automatisch – Prüfung des 1099-B Broker-Auszugs nötig.</li>
        </ul>
        <h3>Steuerformular-Export: Schedule B / 1040 / 1042-S (Letter-Format)</h3>
        <ul>
            <li><b>Auswahl:</b> Investor-Status + Bundesstaat + Filing Status (bei US-Resident)</li>
            <li>Gross income &bull; Withholding &bull; Qualified Div Rates &bull; State Tax &bull; NIIT 3,8 % &bull; Wash Sale Warnbox</li>
        </ul>
        <div class="tip"><b>Tipp:</b> Das PDF zeigt das exakte <b>Formularfeld</b> (z.B. «1042-S Box 7»).</div>
        <div class="warn">&#9888; <b>Orientierungshilfe</b> – offizielle Abgabe via <b>IRS.gov</b>. Bei Unsicherheiten Steuerberater konsultieren.</div>""",

            "EN": _CSS + """
        <h3>&#127482;&#127480; Tax Module USA (US)</h3>
        <p>The US tax module calculates dividends ex-dividend-day-accurately, withholding tax (DTT or 30% standard),
        Qualified Dividend Rates, State Tax, and NIIT.</p>
        <h3>Overview</h3>
        <table>
            <tr><th>Currency</th><th>Tax system</th><th>Key features</th></tr>
            <tr><td>USD</td><td>Federal + State Tax</td><td>DTT CH/DE/AT–US 15% &bull; Standard WHT 30% &bull; Qualified Dividends 0/15/20% &bull; State Tax by state</td></tr>
        </table>
        <h3>Investor Profiles</h3>
        <table>
            <tr><th>Status</th><th>Form</th><th>Withholding tax</th></tr>
            <tr><td><b>CH / DE / AT</b></td><td>1042-S</td><td>15% (DTT)</td></tr>
            <tr><td><b>NRA</b></td><td>1042-S</td><td>30%</td></tr>
            <tr><td><b>US Resident</b></td><td>Form 1040 / Schedule B</td><td>0/15/20% by Filing Status</td></tr>
        </table>
        <h3>Filing Status and Income Thresholds 2024 (US Resident)</h3>
        <table>
            <tr><th>Filing Status</th><th>0% rate up to</th><th>15% rate up to</th><th>20% from</th></tr>
            <tr><td>Single</td><td>$ 47,025</td><td>$ 518,900</td><td>&gt; $ 518,900</td></tr>
            <tr><td>Married Filing Jointly</td><td>$ 94,050</td><td>$ 583,750</td><td>&gt; $ 583,750</td></tr>
            <tr><td>Married Filing Separately</td><td>$ 47,025</td><td>$ 291,850</td><td>&gt; $ 291,850</td></tr>
            <tr><td>Head of Household</td><td>$ 63,000</td><td>$ 551,350</td><td>&gt; $ 551,350</td></tr>
        </table>
        <h3>Wash Sale Rule (IRC § 1091)</h3>
        <ul>
            <li>Applies to <b>US taxpayers only</b>. Swiss, German and Austrian residents are generally <b>not affected</b>.</li>
            <li>The module does not calculate this automatically – check your broker's 1099-B statement.</li>
        </ul>
        <h3>Tax Form Export: Schedule B / 1040 / 1042-S (Letter format)</h3>
        <ul>
            <li><b>Selection:</b> Investor status + State + Filing Status (US Resident only)</li>
            <li>Gross income &bull; Withholding &bull; Qualified Div Rates &bull; State Tax &bull; NIIT 3.8% &bull; Wash Sale warning box</li>
        </ul>
        <div class="tip"><b>Tip:</b> The PDF shows the exact <b>form field</b> (e.g. "1042-S Box 7").</div>
        <div class="warn">&#9888; <b>Guidance tool only</b> – file officially via <b>IRS.gov</b>. Consult a tax advisor if in doubt.</div>""",
        },
    }

    def _show_help():
        from PyQt6.QtWidgets import (
            QDialog as _QD, QVBoxLayout as _QVL, QHBoxLayout as _QHL,
            QTextBrowser as _QTB, QPushButton as _QPB,
            QComboBox as _QCB, QLabel as _QL, QWidget as _QW,
        )
        from PyQt6.QtCore import Qt as _Qt
        from tax_translations import set_tax_language, get_tax_language
        from config import save_config

        _help_dlg = _QD(dlg)
        _help_dlg.setWindowTitle(TRT("tax_help_title", name=rules.get('name', country_code)))
        _help_dlg.setMinimumSize(620, 500)
        _help_dlg.resize(680, 560)
        _help_dlg.setWindowFlags(
            _help_dlg.windowFlags() | _Qt.WindowType.WindowMaximizeButtonHint
        )
        _vl = _QVL(_help_dlg)

        # ── Language switcher row ─────────────────────────────────────────
        _lang_row = _QHL()
        _lang_row.setContentsMargins(0, 0, 0, 4)
        _lang_lbl = _QL(TRT("tax_help_lang_label"))
        _lang_lbl.setStyleSheet("color:#555; font-size:11px;")
        _lang_combo = _QCB()
        _lang_combo.addItem("DE")
        _lang_combo.addItem("EN")
        _lang_combo.setCurrentText(get_tax_language())
        _lang_combo.setMaximumWidth(60)
        _lang_combo.setStyleSheet("font-size:11px;")
        _lang_row.addWidget(_lang_lbl)
        _lang_row.addWidget(_lang_combo)
        _lang_row.addStretch()
        _vl.addLayout(_lang_row)

        # ── Help text browser ─────────────────────────────────────────────
        _tb = _QTB()
        _tb.setOpenExternalLinks(False)
        _tb.setReadOnly(True)

        def _get_html(lang):
            entry = _HELP_TEXTS.get(country_code, {})
            if isinstance(entry, dict):
                return entry.get(lang) or entry.get("DE") or TRT("tax_help_no_text")
            return entry or TRT("tax_help_no_text")   # legacy: plain string

        _tb.setHtml(_get_html(get_tax_language()))
        _vl.addWidget(_tb)

        # ── Language change handler ───────────────────────────────────────
        def _on_lang_changed(new_lang):
            set_tax_language(new_lang)
            save_config({"language": new_lang})
            _help_dlg.setWindowTitle(TRT("tax_help_title", name=rules.get('name', country_code)))
            _lang_lbl.setText(TRT("tax_help_lang_label"))
            _ok.setText(TRT("tax_help_btn_close"))
            _tb.setHtml(_get_html(new_lang))

        _lang_combo.currentTextChanged.connect(_on_lang_changed)

        # ── Close button ──────────────────────────────────────────────────
        _hl = _QHL()
        _hl.addStretch()
        _ok = _QPB(TRT("tax_help_btn_close"))
        _ok.setMinimumHeight(32)
        _ok.clicked.connect(_help_dlg.close)
        _hl.addWidget(_ok)
        _vl.addLayout(_hl)
        _help_dlg.exec()

    help_btn.clicked.connect(_show_help)


    return dlg


# ── Speichern / Laden (AES-256-verschlüsselt, gleiche Logik wie Portfolios) ──

TAX_DIR = os.path.expanduser("~/.stock_monitor_tax")
TAX_EXT = ".smtx"   # stock monitor tax

def _ensure_tax_dir():
    os.makedirs(TAX_DIR, exist_ok=True)

def _save_tax_data(parent_widget, tax_data, country_code):
    from PyQt6.QtWidgets import (
        QFileDialog, QMessageBox, QDialog, QVBoxLayout,
        QLabel, QLineEdit, QHBoxLayout, QPushButton, QWidget, QCheckBox
    )
    from PyQt6.QtCore import Qt
    from PyQt6.QtGui import QFont
    import json, os

    # Dateinamen wählen
    _ensure_tax_dir()
    path, _ = QFileDialog.getSaveFileName(
        parent_widget,
        TRT("tax_save_dialog_title"),
        os.path.join(TAX_DIR, TRT("tax_file_save_default",
            country=country_code, year=tax_data['rules']['year']) + TAX_EXT),
        TRT("tax_file_filter_smtx")
    )
    if not path:
        return

    # Passwort abfragen – mit Stärke-Balken und Sichtbar-Checkbox
    pw_dlg = QDialog(parent_widget)
    pw_dlg.setWindowTitle(TRT("tax_pw_set_title"))
    pw_dlg.setFixedWidth(400)
    lay = QVBoxLayout(pw_dlg)
    lay.setSpacing(8)

    # Icon + Titel
    icon_lbl = QLabel("🔐")
    icon_lbl.setFont(QFont("Segoe UI Emoji", 32))
    icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    lay.addWidget(icon_lbl)
    title_lbl = QLabel(TRT("tax_pw_set_heading"))
    title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    lay.addWidget(title_lbl)
    hint_lbl = QLabel(TRT("tax_pw_set_hint"))
    hint_lbl.setWordWrap(True)
    hint_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    lay.addWidget(hint_lbl)

    lay.addWidget(QLabel(TRT("tax_pw_label")))
    pw1 = QLineEdit(); pw1.setEchoMode(QLineEdit.EchoMode.Password)
    pw1.setPlaceholderText(TRT("tax_pw_placeholder"))
    lay.addWidget(pw1)

    # Stärke-Balken
    bar_container = QWidget(); bar_container.setFixedHeight(14)
    bar_container.setStyleSheet("background:#e0e0e0; border-radius:6px;")
    bar_cl = QHBoxLayout(bar_container)
    bar_cl.setContentsMargins(0,0,0,0); bar_cl.setSpacing(0)
    strength_bar = QWidget(bar_container); strength_bar.setFixedHeight(14)
    strength_bar.setFixedWidth(0)
    strength_bar.setStyleSheet("background:#e0e0e0; border-radius:6px;")
    bar_cl.addWidget(strength_bar); bar_cl.addStretch()
    lay.addWidget(bar_container)
    strength_lbl = QLabel(""); strength_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    strength_lbl.setStyleSheet("font-size:11px; color:#555;")
    lay.addWidget(strength_lbl)

    lay.addWidget(QLabel(TRT("tax_pw_confirm_label")))
    pw2 = QLineEdit(); pw2.setEchoMode(QLineEdit.EchoMode.Password)
    pw2.setPlaceholderText(TRT("tax_pw_confirm_placeholder"))
    lay.addWidget(pw2)

    # Sichtbar-Checkbox
    show_cb = QCheckBox(TRT("tax_pw_show_cb"))
    lay.addWidget(show_cb)

    def _toggle_vis(checked):
        m = QLineEdit.EchoMode.Normal if checked else QLineEdit.EchoMode.Password
        pw1.setEchoMode(m); pw2.setEchoMode(m)
    show_cb.toggled.connect(_toggle_vis)

    def _pw_strength(pw):
        import re
        if not pw: return 0, "", "#e0e0e0"
        score = 0
        n = len(pw)
        if n >= 12: score += 20
        if n >= 16: score += 10
        if n >= 20: score += 20
        score += min(n * 1.5, 20)
        v = sum([bool(re.search(p, pw)) for p in [r'[a-z]',r'[A-Z]',r'\d',r'[^a-zA-Z0-9]']])
        score += v * 10 + (10 if v == 4 else 0)
        score = min(int(score), 100)
        if score < 30:  return score, TRT("tax_pw_strength_weak"),      "#e74c3c"
        if score < 55:  return score, TRT("tax_pw_strength_medium"),    "#e67e22"
        if score < 78:  return score, TRT("tax_pw_strength_strong"),    "#27ae60"
        return score, TRT("tax_pw_strength_very_strong"), "#1a6b3a"

    def _update_strength(text):
        score, label, color = _pw_strength(text)
        total_w = bar_container.width() or 360
        strength_bar.setFixedWidth(max(int(total_w * score / 100), 0))
        strength_bar.setStyleSheet(f"background:{color}; border-radius:6px;")
        strength_lbl.setText(label)
        strength_lbl.setStyleSheet(f"font-size:11px; font-weight:bold; color:{color};")
    pw1.textChanged.connect(_update_strength)

    status_lbl = QLabel("")
    status_lbl.setStyleSheet("color:#e74c3c; font-size:11px;")
    status_lbl.setWordWrap(True)
    lay.addWidget(status_lbl)

    br = QHBoxLayout()
    can = QPushButton(TRT("tax_pw_btn_cancel"))
    ok  = QPushButton(TRT("tax_pw_btn_save"))
    ok.setStyleSheet("background:#27ae60;color:white;font-weight:bold;padding:6px 18px;")
    ok.setDefault(True)
    br.addStretch(); br.addWidget(can); br.addWidget(ok)
    lay.addLayout(br)
    can.clicked.connect(pw_dlg.reject)

    def do_save():
        p1, p2 = pw1.text(), pw2.text()
        if len(p1) < 12:
            status_lbl.setText(TRT("tax_pw_err_short"))
            return
        if p1 != p2:
            status_lbl.setText(TRT("tax_pw_err_mismatch"))
            return
        pw_dlg.accept()
        # Verschlüsseln – AES-256-GCM (gleicher Standard wie Portfolio-Dateien)
        try:
            from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
            from cryptography.hazmat.primitives          import hashes
            from cryptography.hazmat.backends            import default_backend
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            import secrets, base64

            salt  = secrets.token_bytes(32)
            kdf   = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32,
                                salt=salt, iterations=600_000,
                                backend=default_backend())
            key   = kdf.derive(p1.encode())
            nonce = secrets.token_bytes(12)   # 96-bit nonce für GCM

            # Daten serialisieren
            safe = {k: v for k, v in tax_data.items() if k != "rules"}
            raw  = json.dumps(safe, ensure_ascii=False, default=str).encode("utf-8")

            aesgcm = AESGCM(key)
            ct     = aesgcm.encrypt(nonce, raw, None)  # ct enthält Tag (letzten 16 Bytes)

            payload = json.dumps({
                "salt":    base64.b64encode(salt).decode(),
                "nonce":   base64.b64encode(nonce).decode(),
                "ct":      base64.b64encode(ct).decode(),
                "ver":     2,   # v2 = GCM (v1 = CBC legacy)
                "country": country_code,
            })
            with open(path, "w", encoding="utf-8") as f:
                f.write(payload)

            QMessageBox.information(parent_widget, TRT("tax_save_success_title"),
                TRT("tax_save_success_msg", path=path))
        except ImportError:
            # Fallback ohne Verschlüsselung
            with open(path, "w", encoding="utf-8") as f:
                safe = {k: v for k, v in tax_data.items() if k != "rules"}
                json.dump(safe, f, ensure_ascii=False, default=str, indent=2)
            QMessageBox.information(parent_widget, TRT("tax_save_success_title"),
                TRT("tax_save_success_unenc_msg", path=path))
        except Exception as e:
            QMessageBox.warning(parent_widget, TRT("tax_save_error_title"),
                TRT("tax_save_error_msg", error=e))

    ok.clicked.connect(do_save)
    import sys as _sys_sv
    if _sys_sv.platform == 'win32':
        pw_dlg.finished.connect(lambda: (parent_widget.raise_(), parent_widget.activateWindow()) if parent_widget else None)
    pw_dlg.exec()


def _load_tax_data(parent_widget, tax_data_ref, tbl_widget, sum_frame):
    """Lädt einen gespeicherten Steuerauszug und aktualisiert die Anzeige."""
    from PyQt6.QtWidgets import (
        QFileDialog, QMessageBox, QDialog, QVBoxLayout,
        QLabel, QLineEdit, QHBoxLayout, QPushButton, QCheckBox
    )
    from PyQt6.QtCore import Qt
    from PyQt6.QtGui import QFont
    import json, os

    _ensure_tax_dir()
    path, _ = QFileDialog.getOpenFileName(
        parent_widget,
        TRT("tax_load_dialog_title"),
        TAX_DIR,
        f"{TRT('tax_file_filter_smtx')};;{TRT('tax_file_filter_all')}"
    )
    if not path:
        return

    # Datei lesen
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw_content = f.read()
        file_data = json.loads(raw_content)
    except Exception as e:
        QMessageBox.warning(parent_widget, TRT("tax_load_error_title"),
            TRT("tax_load_error_msg", error=e))
        return

    # Prüfen ob verschlüsselt
    if "ct" in file_data:
        pw_dlg = QDialog(parent_widget)
        pw_dlg.setWindowTitle(TRT("tax_pw_enter_title"))
        pw_dlg.setFixedWidth(380)
        lay = QVBoxLayout(pw_dlg)
        lay.setSpacing(8)

        icon_lbl = QLabel("🔐")
        
        icon_lbl.setFont(QFont("Segoe UI Emoji", 32))
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(icon_lbl)
        lay.addWidget(QLabel(TRT("tax_pw_enter_for_file", filename=os.path.basename(path))))

        pw_edit = QLineEdit(); pw_edit.setEchoMode(QLineEdit.EchoMode.Password)
        pw_edit.setPlaceholderText(TRT("tax_pw_enter_placeholder"))
        lay.addWidget(pw_edit)

        
        show_cb2 = QCheckBox(TRT("tax_pw_show_cb"))
        def _tog2(c): pw_edit.setEchoMode(
            QLineEdit.EchoMode.Normal if c else QLineEdit.EchoMode.Password)
        show_cb2.toggled.connect(_tog2)
        lay.addWidget(show_cb2)

        br = QHBoxLayout()
        ok  = QPushButton(TRT("tax_pw_enter_btn_open"))
        ok.setStyleSheet("background:#27ae60;color:white;font-weight:bold;padding:6px 18px;")
        ok.setDefault(True)
        can = QPushButton(TRT("tax_pw_enter_btn_cancel"))
        br.addStretch(); br.addWidget(can); br.addWidget(ok)
        lay.addLayout(br)
        can.clicked.connect(pw_dlg.reject)
        pw_edit.returnPressed.connect(ok.click)

        decrypted = [None]

        def do_load():
            try:
                from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
                from cryptography.hazmat.primitives          import hashes
                from cryptography.hazmat.backends            import default_backend
                import base64

                pw   = pw_edit.text().encode()
                salt = base64.b64decode(file_data["salt"])
                ct   = base64.b64decode(file_data["ct"])
                ver  = file_data.get("ver", 1)

                kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32,
                                  salt=salt, iterations=600_000,
                                  backend=default_backend())
                key = kdf.derive(pw)

                if ver == 2:
                    # AES-256-GCM (neue Dateien)
                    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
                    nonce  = base64.b64decode(file_data["nonce"])
                    aesgcm = AESGCM(key)
                    plain  = aesgcm.decrypt(nonce, ct, None)
                else:
                    # AES-256-CBC (Legacy v1)
                    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
                    iv     = base64.b64decode(file_data["iv"])
                    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
                    dec    = cipher.decryptor()
                    plain  = dec.update(ct) + dec.finalize()
                    pad    = plain[-1]
                    plain  = plain[:-pad]

                decrypted[0] = json.loads(plain.decode("utf-8"))
                pw_dlg.accept()
            except Exception:
                QMessageBox.warning(pw_dlg, TRT("tax_pw_wrong_title"),
                    TRT("tax_pw_wrong_msg"))

        ok.clicked.connect(do_load)
        import sys as _sys_ld
        if _sys_ld.platform == 'win32':
            pw_dlg.finished.connect(lambda: (parent_widget.raise_(), parent_widget.activateWindow()) if parent_widget else None)
        pw_dlg.exec()
        loaded = decrypted[0]
    else:
        loaded = file_data   # unverschlüsselt

    if loaded is None:
        return

    QMessageBox.information(parent_widget, TRT("tax_load_success_title"),
        TRT("tax_load_success_msg", filename=os.path.basename(path)))


# ── Steuerformular-Export ────────────────────────────────────────────────────

def _get_fx_rate_year_end(from_ccy, to_ccy, year):
    """
    Holt Wechselkurs zum 31.12. des Jahres via yfinance.
    Fallback: bekannte Näherungswerte.
    """
    if from_ccy == to_ccy:
        return 1.0
    try:
        import yfinance as yf
        from datetime import date, timedelta
        ticker_sym = f"{from_ccy}{to_ccy}=X"
        end   = date(year, 12, 31)
        start = date(year, 12, 20)
        t = yf.Ticker(ticker_sym)
        hist = t.history(start=start.strftime("%Y-%m-%d"),
                         end=(end + timedelta(days=1)).strftime("%Y-%m-%d"))
        if hist is not None and not hist.empty:
            return float(hist["Close"].iloc[-1])
    except Exception:
        pass
    FALLBACK = {
        ("USD", "CHF"): 0.896, ("CHF", "USD"): 1.116,
        ("EUR", "CHF"): 0.939, ("CHF", "EUR"): 1.065,
        ("GBP", "CHF"): 1.136, ("CHF", "GBP"): 0.880,
        ("USD", "EUR"): 0.924, ("EUR", "USD"): 1.082,
        ("GBP", "EUR"): 1.209, ("EUR", "GBP"): 0.827,
        ("USD", "GBP"): 0.788, ("GBP", "USD"): 1.269,
    }
    return FALLBACK.get((from_ccy, to_ccy), 1.0)


def _export_tax_form(parent_widget, tax_data, country_code):
    """
    Dialog zur Auswahl von Kanton (CH) / Bundesland (DE) / Region (AT)
    und anschliessendem PDF-Export des Steuerformulars.
    """
    from PyQt6.QtWidgets import (
        QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
        QComboBox, QFileDialog, QMessageBox, QFrame, QWidget
    )
    from PyQt6.QtCore import Qt
    from PyQt6.QtGui import QFont, QFontDatabase
    import sys as _sys

    # Emoji-Font ermitteln
    _ef_form = None
    try:
        for _fc in ("Segoe UI Emoji", "Noto Color Emoji", "Apple Color Emoji"):
            if _fc in QFontDatabase.families():
                _ef_form = QFont(_fc, 11)
                break
    except Exception:
        pass

    dlg = QDialog(parent_widget)
    dlg.setWindowTitle(TRT("tax_form_title"))
    dlg.setFixedWidth(480)
    lyt = QVBoxLayout(dlg)
    lyt.setContentsMargins(20, 18, 20, 16)
    lyt.setSpacing(10)

    title_lbl  = QLabel()
    title_lbl.setStyleSheet("font-size:14px; font-weight:bold; color:#1a1a2e;")
    info_lbl   = QLabel()
    info_lbl.setWordWrap(True)
    info_lbl.setStyleSheet("color:#444; font-size:11px;")
    lyt.addWidget(title_lbl)
    lyt.addWidget(info_lbl)

    sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
    sep.setStyleSheet("color:#ddd;"); lyt.addWidget(sep)

    region_lbl   = QLabel()
    region_lbl.setStyleSheet("font-weight:bold; margin-top:4px;")
    region_combo = QComboBox()
    region_combo.setMinimumHeight(28)
    stichtag_lbl = QLabel()
    stichtag_lbl.setStyleSheet("color:#666; font-size:10px;")
    lyt.addWidget(region_lbl)
    lyt.addWidget(region_combo)
    lyt.addWidget(stichtag_lbl)

    _disc_frame = QWidget()
    _disc_frame.setStyleSheet(
        "QWidget { background:#fff8e1; border-left:3px solid #ffc107; border-radius:3px; }"
    )
    _disc_hl = QHBoxLayout(_disc_frame)
    _disc_hl.setContentsMargins(8, 6, 8, 6)
    _disc_hl.setSpacing(8)
    _disc_icon = QLabel("⚠️")
    _disc_icon.setStyleSheet("border:none; background:transparent; font-size:13px;")
    _disc_icon.setAlignment(Qt.AlignmentFlag.AlignTop)
    if _ef_form:
        _disc_icon.setFont(_ef_form)
    _disc_hl.addWidget(_disc_icon, 0, Qt.AlignmentFlag.AlignTop)
    disc_lbl = QLabel()
    disc_lbl.setWordWrap(True)
    disc_lbl.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
    disc_lbl.setStyleSheet("border:none; background:transparent; font-size:10px; color:#555;")
    _disc_hl.addWidget(disc_lbl, 1)
    lyt.addWidget(_disc_frame)

    btn_row    = QHBoxLayout()
    ok_btn     = QPushButton(TRT("tax_form_btn_export"))
    ok_btn.setStyleSheet(
        "background:#1a6b3a; color:white; font-weight:bold; padding:6px 18px;"
    )
    ok_btn.setMinimumHeight(34)
    cancel_btn = QPushButton(TRT("tax_form_btn_cancel"))
    cancel_btn.setMinimumHeight(34)
    btn_row.addWidget(ok_btn)
    btn_row.addWidget(cancel_btn)
    lyt.addLayout(btn_row)

    cancel_btn.clicked.connect(dlg.reject)
    if _sys.platform == "win32":
        dlg.finished.connect(
            lambda: (parent_widget.raise_(), parent_widget.activateWindow())
        )

    year = tax_data["rules"]["year"]
    _do_export = [None]   # Closure-Container

    # ── CH ───────────────────────────────────────────────────────────────────
    if country_code == "CH":
        title_lbl.setText(TRT("tax_form_ch_heading"))
        info_lbl.setText(TRT("tax_form_ch_info"))
        region_lbl.setText(TRT("tax_form_ch_region_label"))
        KANTONE = [
            ("AG","Aargau"), ("AI","Appenzell Innerrhoden"), ("AR","Appenzell Ausserrhoden"),
            ("BE","Bern"), ("BL","Basel-Landschaft"), ("BS","Basel-Stadt"),
            ("FR","Freiburg"), ("GE","Genf"), ("GL","Glarus"),
            ("GR","Graubünden"), ("JU","Jura"), ("LU","Luzern"),
            ("NE","Neuenburg"), ("NW","Nidwalden"), ("OW","Obwalden"),
            ("SG","St. Gallen"), ("SH","Schaffhausen"), ("SO","Solothurn"),
            ("SZ","Schwyz"), ("TG","Thurgau"), ("TI","Tessin"),
            ("UR","Uri"), ("VD","Waadt"), ("VS","Wallis"),
            ("ZG","Zug"), ("ZH","Zürich"),
        ]
        for code, name in KANTONE:
            region_combo.addItem(f"{code}  –  {name}", code)
        # ZH als Standard
        for i, (c, _) in enumerate(KANTONE):
            if c == "ZH":
                region_combo.setCurrentIndex(i); break
        stichtag_lbl.setText(TRT("tax_form_ch_stichtag", year=year))
        disc_lbl.setText(TRT("tax_form_ch_disc", year=year))
        def _do_ch():
            canton_code = region_combo.currentData()
            canton_name = region_combo.currentText().split("–")[-1].strip()
            path, _ = QFileDialog.getSaveFileName(
                dlg, TRT("tax_form_ch_save_dialog"),
                TRT("tax_form_ch_default_filename", canton=canton_code, year=year),
                "PDF (*.pdf)"
            )
            if path:
                try:
                    _export_ch_form_pdf(path, tax_data, canton_code, canton_name, year)
                    QMessageBox.information(dlg, TRT("tax_form_ch_success_title"),
                        TRT("tax_form_ch_success_msg", canton_name=canton_name, path=path))
                    dlg.accept()
                except Exception as e:
                    QMessageBox.critical(dlg, TRT("tax_form_ch_error_title"),
                        TRT("tax_form_ch_error_msg", error=e))
        _do_export[0] = _do_ch

    # ── DE ───────────────────────────────────────────────────────────────────
    elif country_code == "DE":
        title_lbl.setText(TRT("tax_form_de_heading"))
        info_lbl.setText(TRT("tax_form_de_info"))
        region_lbl.setText(TRT("tax_form_de_region_label"))
        BUNDESLAENDER = [
            ("BW","Baden-Württemberg"), ("BY","Bayern"), ("BE","Berlin"),
            ("BB","Brandenburg"), ("HB","Bremen"), ("HH","Hamburg"),
            ("HE","Hessen"), ("MV","Mecklenburg-Vorpommern"), ("NI","Niedersachsen"),
            ("NW","Nordrhein-Westfalen"), ("RP","Rheinland-Pfalz"), ("SL","Saarland"),
            ("SN","Sachsen"), ("ST","Sachsen-Anhalt"), ("SH","Schleswig-Holstein"),
            ("TH","Thüringen"),
        ]
        for code, name in BUNDESLAENDER:
            region_combo.addItem(f"{code}  –  {name}", code)
        region_combo.setCurrentIndex(0)
        stichtag_lbl.setText(TRT("tax_form_de_stichtag", year=year))
        disc_lbl.setText(TRT("tax_form_de_disc", year=year))
        def _do_de():
            bl_code = region_combo.currentData()
            bl_name = region_combo.currentText().split("–")[-1].strip()
            path, _ = QFileDialog.getSaveFileName(
                dlg, TRT("tax_form_de_save_dialog"),
                TRT("tax_form_de_default_filename", year=year),
                "PDF (*.pdf)"
            )
            if path:
                try:
                    _export_de_form_pdf(path, tax_data, bl_code, bl_name, year)
                    QMessageBox.information(dlg, TRT("tax_form_de_success_title"),
                        TRT("tax_form_de_success_msg", bl_name=bl_name, path=path))
                    dlg.accept()
                except Exception as e:
                    QMessageBox.critical(dlg, TRT("tax_form_de_error_title"),
                        TRT("tax_form_de_error_msg", error=e))
        _do_export[0] = _do_de

    # ── AT ───────────────────────────────────────────────────────────────────
    elif country_code == "AT":
        title_lbl.setText(TRT("tax_form_at_heading"))
        info_lbl.setText(TRT("tax_form_at_info"))
        region_lbl.setText(TRT("tax_form_at_region_label"))
        FA_REGIONEN = [
            ("ALLE", TRT("tax_fa_alle")),
            ("W","Wien"), ("NOE","Niederösterreich"), ("OOE","Oberösterreich"),
            ("S","Salzburg"), ("T","Tirol"), ("VBG","Vorarlberg"),
            ("K","Kärnten"), ("ST","Steiermark"), ("BGL","Burgenland"),
        ]
        for code, name in FA_REGIONEN:
            region_combo.addItem(name, code)
        region_combo.setCurrentIndex(0)
        stichtag_lbl.setText(TRT("tax_form_at_stichtag", year=year))
        disc_lbl.setText(TRT("tax_form_at_disc", year=year))
        def _do_at():
            fa_code = region_combo.currentData()
            fa_name = region_combo.currentText()
            path, _ = QFileDialog.getSaveFileName(
                dlg, TRT("tax_form_at_save_dialog"),
                TRT("tax_form_at_default_filename", year=year),
                "PDF (*.pdf)"
            )
            if path:
                try:
                    _export_at_form_pdf(path, tax_data, fa_code, fa_name, year)
                    QMessageBox.information(dlg, TRT("tax_form_at_success_title"),
                        TRT("tax_form_at_success_msg", path=path))
                    dlg.accept()
                except Exception as e:
                    QMessageBox.critical(dlg, TRT("tax_form_at_error_title"),
                        TRT("tax_form_at_error_msg", error=e))
        _do_export[0] = _do_at

    # ── UK ───────────────────────────────────────────────────────────────────
    elif country_code == "UK":
        title_lbl.setText(TRT("tax_form_uk_heading"))
        info_lbl.setText(TRT("tax_form_uk_info"))
        region_lbl.setText(TRT("tax_form_uk_region_label"))
        UK_NATIONS = [
            ("ENG", "England"),
            ("WAL", "Wales"),
            ("SCO", "Scotland"),
            ("NIR", "Northern Ireland"),
        ]
        for code, name in UK_NATIONS:
            region_combo.addItem(name, code)
        region_combo.setCurrentIndex(0)
        stichtag_lbl.setText(TRT("tax_form_uk_stichtag", year=year))
        disc_lbl.setText(TRT("tax_form_uk_disc", year=year,
            year_plus1=(year + 1) if isinstance(year, int) else year))
        def _do_uk():
            nation_code = region_combo.currentData()
            nation_name = region_combo.currentText()
            path, _ = QFileDialog.getSaveFileName(
                dlg, TRT("tax_form_uk_save_dialog"),
                TRT("tax_form_uk_default_filename", year=year),
                "PDF (*.pdf)"
            )
            if path:
                try:
                    _export_uk_form_pdf(path, tax_data, nation_code, nation_name, year)
                    QMessageBox.information(dlg, TRT("tax_form_uk_success_title"),
                        TRT("tax_form_uk_success_msg", nation_name=nation_name, path=path))
                    dlg.accept()
                except Exception as e:
                    QMessageBox.critical(dlg, TRT("tax_form_uk_error_title"),
                        TRT("tax_form_uk_error_msg", error=e))
        _do_export[0] = _do_uk

    # ── US ───────────────────────────────────────────────────────────────────
    elif country_code == "US":
        title_lbl.setText(TRT("tax_form_us_heading"))
        info_lbl.setText(TRT("tax_form_us_info"))
        region_lbl.setText(TRT("tax_form_us_region_label"))
        US_TYPES = [
            ("CH",  TRT("tax_us_type_ch")),
            ("DE",  TRT("tax_us_type_de")),
            ("AT",  TRT("tax_us_type_at")),
            ("NRA", TRT("tax_us_type_nra")),
            ("RES", TRT("tax_us_type_res")),
        ]
        for code, name in US_TYPES:
            region_combo.addItem(name, code)
        region_combo.setCurrentIndex(0)
        stichtag_lbl.setText(TRT("tax_form_us_stichtag", year=year))
        disc_lbl.setText(TRT("tax_form_us_disc"))

        # ── State Tax Dropdown (nur für RES sichtbar) ─────────────────────
        US_STATES = [
            # ── Kein State Tax ───────────────────────────────────────────────
            ("NONE", "No State Tax – AK, FL, NV, SD, TX, WA, WY (0 %)"),
            ("NH",   "New Hampshire – 3.0 % (dividends/interest only, ends 2027)"),
            ("TN",   "Tennessee – 0 % (Hall Tax repealed 2021)"),
            # ── Hohe Steuern ─────────────────────────────────────────────────
            ("CA",   "California – 9.3 % / 13.3 % (top)"),
            ("DC",   "Washington D.C. – 8.5 % / 10.75 % (top)"),
            ("NJ",   "New Jersey – 6.37 % / 10.75 % (top)"),
            ("OR",   "Oregon – 8.75 % / 9.9 % (top)"),
            ("MN",   "Minnesota – 7.85 % / 9.85 % (top)"),
            ("NY",   "New York – 6.85 % / 10.9 % (top)"),
            ("VT",   "Vermont – 6.6 % / 8.75 % (top)"),
            ("HI",   "Hawaii – 8.25 % / 11.0 % (top)"),
            ("CT",   "Connecticut – 5.0 % / 6.99 % (top)"),
            ("WI",   "Wisconsin – 5.3 % / 7.65 % (top)"),
            ("SC",   "South Carolina – 3.0 % / 6.5 % (top)"),
            ("ME",   "Maine – 5.8 % / 7.15 % (top)"),
            ("ID",   "Idaho – 5.695 % (flat)"),
            ("MT",   "Montana – 5.9 % / 6.75 % (top)"),
            ("AR",   "Arkansas – 2.0 % / 4.9 % (top)"),
            ("NM",   "New Mexico – 1.7 % / 5.9 % (top)"),
            ("DE",   "Delaware – 2.2 % / 6.6 % (top)"),
            ("NE",   "Nebraska – 2.46 % / 5.84 % (top)"),
            ("WV",   "West Virginia – 3.0 % / 6.5 % (top)"),
            ("RI",   "Rhode Island – 3.75 % / 5.99 % (top)"),
            ("MD",   "Maryland – 2.0 % / 5.75 % (top) + county"),
            # ── Mittlere Steuern ─────────────────────────────────────────────
            ("VA",   "Virginia – 2.0 % / 5.75 % (top)"),
            ("MA",   "Massachusetts – 5.0 % (flat)"),
            ("IL",   "Illinois – 4.95 % (flat)"),
            ("KY",   "Kentucky – 4.0 % (flat)"),
            ("LA",   "Louisiana – 1.85 % / 4.25 % (top)"),
            ("MO",   "Missouri – 1.5 % / 4.95 % (top)"),
            ("AL",   "Alabama – 2.0 % / 5.0 % (top)"),
            ("OH",   "Ohio – 2.765 % / 3.99 % (top)"),
            ("KS",   "Kansas – 3.1 % / 5.7 % (top)"),
            ("OK",   "Oklahoma – 0.25 % / 4.75 % (top)"),
            ("MS",   "Mississippi – 4.7 % (flat)"),
            ("IA",   "Iowa – 3.8 % / 6.0 % (top)"),
            ("GA",   "Georgia – 5.49 % (flat)"),
            ("NC",   "North Carolina – 4.75 % (flat)"),
            ("MI",   "Michigan – 4.05 % (flat)"),
            ("PA",   "Pennsylvania – 3.07 % (flat)"),
            # ── Niedrige Steuern ─────────────────────────────────────────────
            ("UT",   "Utah – 4.65 % (flat)"),
            ("IN",   "Indiana – 3.05 % (flat)"),
            ("CO",   "Colorado – 4.4 % (flat)"),
            ("AZ",   "Arizona – 2.5 % (flat)"),
            ("ND",   "North Dakota – 1.1 % / 2.5 % (top)"),
            ("WY_note", "Wyoming – 0 % (no income tax)"),
            # ── Übrige ──────────────────────────────────────────────────────
            ("WA_note", "Washington State – 0 % (no income tax; capital gains 7 %)"),
            ("SD_note", "South Dakota – 0 % (no income tax)"),
            ("AK_note", "Alaska – 0 % (no income tax)"),
            ("FL_note", "Florida – 0 % (no income tax)"),
            ("TX_note", "Texas – 0 % (no income tax)"),
            ("NV_note", "Nevada – 0 % (no income tax)"),
            ("OTHER",   "Other / Unknown – use 5.0 % as estimate"),
        ]
        state_lbl = QLabel(TRT("tax_form_us_state_label"))
        state_lbl.setStyleSheet("font-weight:bold; margin-top:4px;")
        state_combo = QComboBox()
        state_combo.setMinimumHeight(28)
        for sc, sn in US_STATES:
            state_combo.addItem(sn, sc)
        state_combo.setCurrentIndex(0)
        state_note = QLabel(TRT("tax_form_us_state_note"))
        state_note.setStyleSheet("color:#666; font-size:10px;")
        lyt.addWidget(state_lbl)
        lyt.addWidget(state_combo)
        lyt.addWidget(state_note)

        # ── Filing Status Dropdown (nur für RES sichtbar) ─────────────────
        FILING_STATUSES = [
            ("SINGLE",   "Single"),
            ("MFJ",      "Married Filing Jointly (MFJ)"),
            ("MFS",      "Married Filing Separately (MFS)"),
            ("HOH",      "Head of Household (HOH)"),
        ]
        filing_lbl = QLabel(TRT("tax_form_us_filing_label"))
        filing_lbl.setStyleSheet("font-weight:bold; margin-top:4px;")
        filing_combo = QComboBox()
        filing_combo.setMinimumHeight(28)
        for fc, fn in FILING_STATUSES:
            filing_combo.addItem(fn, fc)
        filing_combo.setCurrentIndex(0)
        filing_note = QLabel(TRT("tax_form_us_filing_note"))
        filing_note.setStyleSheet("color:#666; font-size:10px;")
        lyt.addWidget(filing_lbl)
        lyt.addWidget(filing_combo)
        lyt.addWidget(filing_note)

        def _toggle_state_widgets(idx):
            show = (region_combo.itemData(idx) == "RES")
            state_lbl.setVisible(show)
            state_combo.setVisible(show)
            state_note.setVisible(show)
            filing_lbl.setVisible(show)
            filing_combo.setVisible(show)
            filing_note.setVisible(show)
            dlg.adjustSize()

        # initial hide
        state_lbl.setVisible(False)
        state_combo.setVisible(False)
        state_note.setVisible(False)
        filing_lbl.setVisible(False)
        filing_combo.setVisible(False)
        filing_note.setVisible(False)
        region_combo.currentIndexChanged.connect(_toggle_state_widgets)

        def _do_us():
            inv_code    = region_combo.currentData()
            inv_name    = region_combo.currentText()
            state_code  = state_combo.currentData()  if inv_code == "RES" else None
            state_name  = state_combo.currentText()  if inv_code == "RES" else None
            filing_code = filing_combo.currentData() if inv_code == "RES" else None
            filing_name = filing_combo.currentText() if inv_code == "RES" else None
            path, _ = QFileDialog.getSaveFileName(
                dlg, TRT("tax_form_us_save_dialog"),
                TRT("tax_form_us_default_filename", year=year),
                "PDF (*.pdf)"
            )
            if path:
                try:
                    _export_us_form_pdf(path, tax_data, inv_code, inv_name, year,
                                        state_code=state_code, state_name=state_name,
                                        filing_code=filing_code, filing_name=filing_name)
                    QMessageBox.information(dlg, TRT("tax_form_us_success_title"),
                        TRT("tax_form_us_success_msg", path=path))
                    dlg.accept()
                except Exception as e:
                    QMessageBox.critical(dlg, TRT("tax_form_us_error_title"),
                        TRT("tax_form_us_error_msg", error=e))
        _do_export[0] = _do_us

    ok_btn.clicked.connect(lambda: _do_export[0]() if _do_export[0] else None)
    dlg.exec()


def _export_ch_form_pdf(path, tax_data, canton_code, canton_name, year):
    """Wertschriftenverzeichnis CH (Formular DA-1) – Orientierungshilfe."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    rows    = tax_data["rows"]
    rules   = tax_data["rules"]

    fx_usd = _get_fx_rate_year_end("USD", "CHF", year)
    fx_eur = _get_fx_rate_year_end("EUR", "CHF", year)
    fx_gbp = _get_fx_rate_year_end("GBP", "CHF", year)

    doc = SimpleDocTemplate(path, pagesize=A4,
                            rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm,  bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    DARK  = colors.HexColor("#1a2a6c")
    MID   = colors.HexColor("#2980b9")
    LIGHT = colors.HexColor("#d6eaf8")
    LGREY = colors.HexColor("#f5f5f5")
    RED   = colors.HexColor("#c0392b")
    YELL  = colors.HexColor("#fff9c4")

    def P(txt, fs=8.5, bold=False, color=None, align=0):
        col = color or colors.black
        fn  = "Helvetica-Bold" if bold else "Helvetica"
        return Paragraph(
            f"<font name='{fn}' size='{fs}' color='{col.hexval() if hasattr(col,'hexval') else col}'>{txt}</font>",
            ParagraphStyle("p", parent=styles["Normal"], fontSize=fs,
                           fontName=fn, alignment=align, leading=fs*1.3)
        )

    def chf(v):
        try:
            x = float(v)
            s = f"{abs(x):,.2f}".replace(",", "'")
            return f"-{s}" if x < 0 else s
        except:
            return "–"

    def to_chf(val, ccy):
        try:
            v = float(val)
            if ccy == "CHF": return v
            if ccy == "EUR": return v * fx_eur
            if ccy == "GBP": return v * fx_gbp
            return v * fx_usd
        except:
            return 0.0

    # ── Header ───────────────────────────────────────────────────────────────
    hdr = Table([[
        P(f"Wertschriften- und Guthabenverzeichnis", fs=14, bold=True),
        P(f"DA-1 / R-US\nOrientierungshilfe", fs=8, align=2)
    ]], colWidths=[13*cm, 4*cm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), LIGHT),
        ("BOX",           (0,0),(-1,-1), 0.5, MID),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"<font size='9' color='#2980b9'>"
        f"Kanton {canton_code} – {canton_name} &nbsp;|&nbsp; "
        f"Steuerjahr {year} &nbsp;|&nbsp; Stichtag 31. Dezember {year}"
        f"</font>",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, spaceAfter=4)
    ))

    # ── Steuerpflichtige Person ───────────────────────────────────────────────
    story.append(Paragraph(
        "<b>Angaben zur steuerpflichtigen Person</b>",
        ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                       fontName="Helvetica-Bold", textColor=DARK,
                       spaceBefore=8, spaceAfter=4)
    ))
    pd_tbl = Table([
        ["Name / Vorname:", "____________________________",
         "AHV-Nr.:", "756.____.____.____.__"],
        ["Adresse / Wohnort:", "____________________________",
         "Veranlagungsjahr:", str(year)],
    ], colWidths=[3.5*cm, 6*cm, 2.8*cm, 4.7*cm])
    pd_tbl.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8.5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("LINEBELOW",     (1,0),(1,0),   0.3, colors.grey),
        ("LINEBELOW",     (1,1),(1,1),   0.3, colors.grey),
        ("LINEBELOW",     (3,0),(3,0),   0.3, colors.grey),
    ]))
    story.append(pd_tbl)
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID))
    story.append(Spacer(1, 3))

    # Wechselkurse
    story.append(Paragraph(
        f"<font size='8'><b>Umrechnungskurse per 31.12.{year} (SNB / Yahoo Finance):</b> &nbsp;"
        f"1 USD = CHF {fx_usd:.4f} &nbsp;|&nbsp; "
        f"1 EUR = CHF {fx_eur:.4f} &nbsp;|&nbsp; "
        f"1 GBP = CHF {fx_gbp:.4f}</font>",
        ParagraphStyle("fx", parent=styles["Normal"], fontSize=8, spaceAfter=6,
                       textColor=colors.HexColor("#444"))
    ))

    # ── Teil A: Wertschriften ─────────────────────────────────────────────────
    story.append(Paragraph(
        "<b>Teil A – Wertschriften (Aktien, ETFs)</b>",
        ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                       fontName="Helvetica-Bold", textColor=DARK,
                       spaceBefore=6, spaceAfter=3)
    ))

    HDR = ["Pos.", "Titel", "Stück",
           f"Kurs\n31.12.{year}", "Währ.",
           "Kurswert CHF\n31.12.",
           "Dividenden\nCHF (brutto)",
           "VSt 35%\nCHF",
           "Ausl. QSt\nCHF",
           "Hinweis"]
    tbl_data = [HDR]

    tot_kw = tot_div = tot_vst = tot_qst = 0.0

    for i, row in enumerate(rows, 1):
        ticker = row.get("ticker", "")
        name   = str(row.get("name", ticker))[:26]
        qty    = row.get("qty", row.get("quantity", 0))
        price  = row.get("price", row.get("current_price", 0))
        ccy    = row.get("currency", "USD")
        div    = row.get("div_total", 0)
        vst    = row.get("vst", row.get("withholding", 0))
        qst    = row.get("foreign_tax", 0)
        note   = str(row.get("note", row.get("rate_note", "")))[:28]

        kw_chf  = to_chf(float(qty) * float(price), ccy)
        div_chf = to_chf(div, ccy)
        vst_chf = to_chf(vst, ccy)
        qst_chf = to_chf(qst, ccy)
        tot_kw  += kw_chf; tot_div += div_chf
        tot_vst += vst_chf; tot_qst += qst_chf

        def _n(v, d=2):
            try: return f"{float(v):,.{d}f}".replace(",","'")
            except: return "–"

        tbl_data.append([
            str(i),
            Paragraph(f"<b>{ticker}</b><br/><font size='7' color='#555'>{name}</font>",
                      ParagraphStyle("tc", parent=styles["Normal"], fontSize=8, leading=10)),
            _n(qty, 4 if float(qty) < 100 else 2),
            _n(price), ccy,
            chf(kw_chf),
            chf(div_chf) if div_chf > 0 else "–",
            chf(vst_chf) if vst_chf > 0 else "–",
            chf(qst_chf) if qst_chf > 0 else "–",
            Paragraph(f"<font size='7'>{note}</font>",
                      ParagraphStyle("nt", parent=styles["Normal"], fontSize=7)),
        ])

    # Summe
    def _Pb(txt):
        return Paragraph(f"<b>{txt}</b>",
                         ParagraphStyle("b", parent=styles["Normal"],
                                        fontSize=8.5, fontName="Helvetica-Bold"))
    tbl_data.append(["", _Pb("TOTAL"), "", "", "",
                     _Pb(chf(tot_kw)), _Pb(chf(tot_div)),
                     _Pb(chf(tot_vst)), _Pb(chf(tot_qst)), ""])

    cw = [0.7*cm, 4.0*cm, 1.3*cm, 1.3*cm, 1.0*cm,
          2.2*cm, 2.0*cm, 1.7*cm, 1.7*cm, 2.1*cm]
    wt = Table(tbl_data, colWidths=cw, repeatRows=1)
    wt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  7.5),
        ("ALIGN",         (0,0),(-1,0),  "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("ALIGN",         (2,1),(-2,-1), "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGREY]),
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT),
        ("LINEABOVE",     (0,-1),(-1,-1), 1.0, DARK),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
    ]))
    story.append(wt)
    story.append(Spacer(1, 8))

    # ── Übertrag ─────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.8, color=DARK))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "<b>Übertrag in die Steuererklärung</b>",
        ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                       fontName="Helvetica-Bold", textColor=DARK, spaceAfter=4)
    ))

    sum_rows = [
        ["Feld", "Beschreibung", "CHF-Betrag", "Wo eintragen"],
        ["1", "Gesamtwert Wertschriften per 31.12.",
         _Pb(chf(tot_kw)),
         Paragraph("<font color='#2980b9'>Vermögenssteuer-Formular\nKanton " + canton_code + "</font>",
                   ParagraphStyle("s", parent=styles["Normal"], fontSize=8))],
        ["2", "Erhaltene Dividenden (brutto)",
         _Pb(chf(tot_div)),
         Paragraph("<font color='#2980b9'>Einkommenssteuer\n(Wertschriftenerträge)</font>",
                   ParagraphStyle("s", parent=styles["Normal"], fontSize=8))],
        ["3", "Verrechnungssteuer 35 % (rückforderbar)",
         _Pb(chf(tot_vst)),
         Paragraph("<font color='#c0392b'>Rückforderungsantrag R-CH\n(Formular 25)</font>",
                   ParagraphStyle("s", parent=styles["Normal"], fontSize=8))],
        ["4", "Ausländische Quellensteuer (anrechenbar)",
         _Pb(chf(tot_qst)),
         Paragraph("<font color='#c0392b'>Formular DA-1\n(Antrag auf Steueranrechnung)</font>",
                   ParagraphStyle("s", parent=styles["Normal"], fontSize=8))],
    ]
    st = Table(sum_rows, colWidths=[1.2*cm, 6.8*cm, 3.2*cm, 5.8*cm])
    st.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
        ("FONTSIZE",      (0,1),(-1,-1), 8.5),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("ALIGN",         (2,0),(2,-1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, YELL]),
    ]))
    story.append(st)
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey))
    story.append(Spacer(1, 3))
    story.append(Paragraph(
        f"⚠️  <b>Orientierungshilfe – kein offizielles Steuerformular.</b> "
        f"Bitte verwenden Sie den originalen Bankauszug (Depotauszug per 31.12.{year}) "
        f"und das Online-Steuerformular Ihres Kantons ({canton_name}). "
        f"Abweichungen von ca. 1–3 % möglich. "
        f"Erstellt mit Stock Monitor v5.0.",
        ParagraphStyle("w", parent=styles["Normal"], fontSize=8,
                       textColor=RED, spaceAfter=0)
    ))
    doc.build(story)


def _export_de_form_pdf(path, tax_data, bl_code, bl_name, year):
    """Anlage KAP (Kapitalerträge) DE – Orientierungshilfe."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    rows  = tax_data["rows"]
    rules = tax_data["rules"]

    fx_usd = _get_fx_rate_year_end("USD", "EUR", year)
    fx_chf = _get_fx_rate_year_end("CHF", "EUR", year)
    fx_gbp = _get_fx_rate_year_end("GBP", "EUR", year)

    doc = SimpleDocTemplate(path, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    DARK  = colors.HexColor("#1a2a6c")
    MID   = colors.HexColor("#2471a3")
    LIGHT = colors.HexColor("#d6eaf8")
    LGREY = colors.HexColor("#f5f5f5")
    RED   = colors.HexColor("#c0392b")
    YELL  = colors.HexColor("#fff9c4")

    def eur(v):
        try:
            x = float(v)
            s = f"{abs(x):,.2f}".replace(",","X").replace(".","," ).replace("X",".")
            return f"-{s}" if x < 0 else s
        except: return "–"

    def _n(v, d=2):
        try: return f"{float(v):,.{d}f}".replace(",","X").replace(".","," ).replace("X",".")
        except: return "–"

    def to_eur(val, ccy):
        try:
            v = float(val)
            if ccy == "EUR": return v
            if ccy == "CHF": return v * fx_chf
            if ccy == "GBP": return v * fx_gbp
            return v * fx_usd
        except: return 0.0

    def _Pb(txt):
        return Paragraph(f"<b>{txt}</b>",
                         ParagraphStyle("b", parent=styles["Normal"],
                                        fontSize=8.5, fontName="Helvetica-Bold"))

    # Header
    hdr = Table([[
        Paragraph("<b>Anlage KAP – Einkünfte aus Kapitalvermögen</b>",
                  ParagraphStyle("h1", parent=styles["Normal"], fontSize=13,
                                 fontName="Helvetica-Bold", textColor=DARK)),
        Paragraph(f"Orientierungshilfe\nSteuerjahr {year}",
                  ParagraphStyle("r", parent=styles["Normal"], fontSize=8,
                                 alignment=2, textColor=colors.HexColor("#888")))
    ]], colWidths=[12*cm, 5*cm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), LIGHT),
        ("BOX",           (0,0),(-1,-1), 0.5, MID),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    story.append(hdr)
    story.append(Paragraph(
        f"<font size='9' color='#2471a3'>Bundesland: {bl_name} &nbsp;|&nbsp; "
        f"Steuerjahr {year} &nbsp;|&nbsp; Stichtag 31. Dezember {year}</font>",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, spaceAfter=6,
                       spaceBefore=3)
    ))

    # Persönliche Daten
    story.append(Paragraph("<b>Angaben zur steuerpflichtigen Person</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))
    pt = Table([
        ["Name, Vorname:", "____________________________",
         "Steuernummer:", "____________________________"],
        ["Finanzamt:", f"Finanzamt {bl_name}",
         "Veranlagungszeitraum:", str(year)],
    ], colWidths=[3.5*cm, 6*cm, 3.5*cm, 4*cm])
    pt.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8.5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LINEBELOW",     (1,0),(1,0),   0.4, colors.grey),
        ("LINEBELOW",     (3,0),(3,0),   0.4, colors.grey),
    ]))
    story.append(pt)
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID))
    story.append(Paragraph(
        f"<font size='8'><b>Umrechnungskurse per 31.12.{year} (Bundesbank / Yahoo Finance):</b> "
        f"1 USD = {_n(fx_usd,4)} EUR &nbsp;|&nbsp; "
        f"1 CHF = {_n(fx_chf,4)} EUR &nbsp;|&nbsp; "
        f"1 GBP = {_n(fx_gbp,4)} EUR</font>",
        ParagraphStyle("fx", parent=styles["Normal"], fontSize=8,
                       spaceBefore=4, spaceAfter=8,
                       textColor=colors.HexColor("#444"))
    ))

    # Wertpapier-Tabelle
    story.append(Paragraph("<b>Kapitalerträge – Einzelpositionen (Zeile 7 Anlage KAP)</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=3)))
    HDR = ["Pos.", "Wertpapier", "Stück", f"Kurs\n31.12.", "Währ.",
           "Marktwert EUR", "Dividenden EUR", "Abgeltungst.", "Hinweis"]
    tbl_data = [HDR]

    tot_mw = tot_div = tot_ast = 0.0
    fsa = float(rules.get("freistellungsauftrag", 1000))

    for i, row in enumerate(rows, 1):
        ticker = row.get("ticker","")
        name   = str(row.get("name", ticker))[:24]
        qty    = row.get("qty", row.get("quantity", 0))
        price  = row.get("price", row.get("current_price", 0))
        ccy    = row.get("currency","USD")
        div    = row.get("div_total", 0)
        ast    = row.get("withholding", 0)
        note   = str(row.get("note", row.get("rate_note","")))[:26]

        mw_e  = to_eur(float(qty)*float(price), ccy)
        div_e = to_eur(div, ccy)
        ast_e = to_eur(ast, ccy)
        tot_mw += mw_e; tot_div += div_e; tot_ast += ast_e

        tbl_data.append([
            str(i),
            Paragraph(f"<b>{ticker}</b><br/><font size='7'>{name}</font>",
                      ParagraphStyle("tc", parent=styles["Normal"], fontSize=8, leading=10)),
            _n(qty, 2), _n(price), ccy,
            eur(mw_e),
            eur(div_e) if div_e > 0 else "–",
            eur(ast_e) if ast_e > 0 else "–",
            Paragraph(f"<font size='7'>{note}</font>",
                      ParagraphStyle("nt", parent=styles["Normal"], fontSize=7)),
        ])

    tbl_data.append(["", _Pb("TOTAL"), "", "", "",
                     _Pb(eur(tot_mw)), _Pb(eur(tot_div)), _Pb(eur(tot_ast)), ""])

    cw = [0.7*cm, 4.2*cm, 1.3*cm, 1.3*cm, 1.0*cm,
          2.4*cm, 2.2*cm, 2.0*cm, 2.1*cm]
    kt = Table(tbl_data, colWidths=cw, repeatRows=1)
    kt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  7.5),
        ("ALIGN",         (0,0),(-1,0),  "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("ALIGN",         (2,1),(-2,-1), "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGREY]),
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT),
        ("LINEABOVE",     (0,-1),(-1,-1), 1.0, DARK),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
    ]))
    story.append(kt)
    story.append(Spacer(1, 8))

    # Anlage KAP Übertrag
    story.append(HRFlowable(width="100%", thickness=0.8, color=DARK))
    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>Übertrag in die Anlage KAP</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))

    div_nach_fsa = max(0.0, tot_div - fsa)
    abgst        = div_nach_fsa * rules.get("withholding_rate", 0.25)
    soli         = abgst * rules.get("soli_rate", 0.055)
    gesamt       = abgst + soli

    def _Pe(txt):
        """Erläuterungs-Paragraph mit Zeilenumbruch."""
        return Paragraph(txt, ParagraphStyle("e", parent=styles["Normal"],
                                              fontSize=8, leading=10,
                                              textColor=colors.HexColor("#444")))

    kap_rows = [
        ["Zeile", "Bezeichnung", "Betrag EUR", "Erläuterung"],
        ["7",  "Kapitalerträge gesamt (Dividenden)",
         _Pb(eur(tot_div)), _Pe(f"Summe Brutto-Dividenden (Steuerjahr {year})")],
        ["–",  "– Sparer-Pauschbetrag (Freistellungsauftrag)",
         _Pb(f"– {eur(fsa)}"), _Pe(f"Max. {eur(fsa)} EUR p.P.\n(§ 20 Abs. 9 EStG)")],
        ["–",  "Steuerpflichtige Kapitalerträge",
         Paragraph(f"<b><font color='#c0392b'>{eur(div_nach_fsa)}</font></b>",
                   ParagraphStyle("r", parent=styles["Normal"],
                                  fontSize=9, fontName="Helvetica-Bold")),
         _Pe("Basis für Abgeltungsteuer")],
        ["–",  "Abgeltungsteuer 25 %",
         _Pb(eur(abgst)), _Pe("25 % auf steuerpfl. Kapitalerträge")],
        ["–",  "Solidaritätszuschlag 5,5 %",
         _Pb(eur(soli)), _Pe("5,5 % auf Abgeltungsteuer")],
        ["–",  "Steuerbelastung gesamt (Schätzung)",
         Paragraph(f"<b><font color='#c0392b'>{eur(gesamt)}</font></b>",
                   ParagraphStyle("r", parent=styles["Normal"],
                                  fontSize=10, fontName="Helvetica-Bold")),
         _Pe(f"Ohne Kirchensteuer\n({bl_name})")],
        ["–",  "Bereits einbehaltene QSt (anrechenbar)",
         _Pb(eur(tot_ast)), _Pe("Anrechenbar auf Abgeltungsteuer")],
        ["62", "Marktwert Wertschriften per 31.12.",
         _Pb(eur(tot_mw)), _Pe("Bestandsnachweis\n(Anlage KAP Zeile 62)")],
    ]
    # Spalten: Zeile | Bezeichnung | Betrag | Erläuterung  → total 17cm (A4 - 4cm Ränder)
    kt2 = Table(kap_rows, colWidths=[1.2*cm, 7.0*cm, 3.3*cm, 5.5*cm])
    kt2.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
        ("FONTSIZE",      (0,1),(-1,-1), 8.5),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("ALIGN",         (2,0),(2,-1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, YELL]),
    ]))
    story.append(kt2)
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey))
    story.append(Spacer(1, 3))
    story.append(Paragraph(
        f"⚠️  <b>Orientierungshilfe – kein offizielles ELSTER-Formular.</b> "
        f"Bitte verwenden Sie die Jahressteuerbescheinigung Ihrer Depotbank "
        f"sowie ELSTER (elster.de). Kirchensteuer ({bl_name}) nicht berücksichtigt. "
        f"Abweichungen von ca. 1–3 % möglich. Erstellt mit Stock Monitor v5.0.",
        ParagraphStyle("w", parent=styles["Normal"], fontSize=8, textColor=RED)
    ))
    doc.build(story)


def _export_at_form_pdf(path, tax_data, fa_code, fa_name, year):
    """E1kv-Nachweis AT (Kapitalerträge-Beilage) – Orientierungshilfe."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    rows  = tax_data["rows"]
    rules = tax_data["rules"]

    fx_usd = _get_fx_rate_year_end("USD", "EUR", year)
    fx_chf = _get_fx_rate_year_end("CHF", "EUR", year)
    fx_gbp = _get_fx_rate_year_end("GBP", "EUR", year)

    doc = SimpleDocTemplate(path, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    DARK  = colors.HexColor("#8b0000")
    MID   = colors.HexColor("#c0392b")
    LIGHT = colors.HexColor("#fde8e8")
    LGREY = colors.HexColor("#f5f5f5")
    YELL  = colors.HexColor("#fff9c4")

    def eur(v):
        try:
            x = float(v)
            s = f"{abs(x):,.2f}".replace(",","X").replace(".","," ).replace("X",".")
            return f"-{s}" if x < 0 else s
        except: return "–"

    def _n(v, d=2):
        try: return f"{float(v):,.{d}f}".replace(",","X").replace(".","," ).replace("X",".")
        except: return "–"

    def to_eur(val, ccy):
        try:
            v = float(val)
            if ccy == "EUR": return v
            if ccy == "CHF": return v * fx_chf
            if ccy == "GBP": return v * fx_gbp
            return v * fx_usd
        except: return 0.0

    def _Pb(txt):
        return Paragraph(f"<b>{txt}</b>",
                         ParagraphStyle("b", parent=styles["Normal"],
                                        fontSize=8.5, fontName="Helvetica-Bold"))

    # Header
    fa_hint = f" &nbsp;|&nbsp; FA-Region: {fa_name}" if fa_code != "ALLE" else ""
    hdr = Table([[
        Paragraph("<b>Nachweis Kapitalerträge – Beilage E1kv</b>",
                  ParagraphStyle("h1", parent=styles["Normal"], fontSize=13,
                                 fontName="Helvetica-Bold", textColor=DARK)),
        Paragraph(f"Orientierungshilfe\nSteuerjahr {year}",
                  ParagraphStyle("r", parent=styles["Normal"], fontSize=8,
                                 alignment=2, textColor=colors.HexColor("#888")))
    ]], colWidths=[12*cm, 5*cm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), LIGHT),
        ("BOX",           (0,0),(-1,-1), 0.5, MID),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    story.append(hdr)
    story.append(Paragraph(
        f"<font size='9' color='#c0392b'>Österreich &nbsp;|&nbsp; "
        f"Steuerjahr {year} &nbsp;|&nbsp; Stichtag 31. Dezember {year}"
        f"{fa_hint}</font>",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
                       spaceAfter=6, spaceBefore=3)
    ))

    # Persönliche Daten
    story.append(Paragraph("<b>Angaben zur steuerpflichtigen Person</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))
    pt = Table([
        ["Name, Vorname:", "____________________________",
         "Sozialversicherungsnr.:", "____________________________"],
        ["Adresse:", "____________________________",
         "Veranlagungsjahr:", str(year)],
    ], colWidths=[3.5*cm, 5.5*cm, 4*cm, 4*cm])
    pt.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8.5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LINEBELOW",     (1,0),(1,0),   0.4, colors.grey),
        ("LINEBELOW",     (3,0),(3,0),   0.4, colors.grey),
        ("LINEBELOW",     (1,1),(1,1),   0.4, colors.grey),
    ]))
    story.append(pt)
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID))
    story.append(Paragraph(
        f"<font size='8'><b>Umrechnungskurse per 31.12.{year} (OeNB / Yahoo Finance):</b> "
        f"1 USD = {_n(fx_usd,4)} EUR &nbsp;|&nbsp; "
        f"1 CHF = {_n(fx_chf,4)} EUR &nbsp;|&nbsp; "
        f"1 GBP = {_n(fx_gbp,4)} EUR</font>",
        ParagraphStyle("fx", parent=styles["Normal"], fontSize=8,
                       spaceBefore=4, spaceAfter=8,
                       textColor=colors.HexColor("#444"))
    ))

    # Wertpapier-Tabelle
    story.append(Paragraph("<b>Aufstellung Kapitalerträge – Einzelpositionen</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=3)))
    HDR = ["Pos.", "Wertpapier", "Stück", f"Kurs\n31.12.", "Währ.",
           "Depotwert EUR", "Dividenden EUR", "KESt 27,5%", "Ausl. QSt EUR", "Hinweis"]
    tbl_data = [HDR]

    tot_depot = tot_div = tot_kest = tot_qst = 0.0
    kest_rate = rules.get("kest_rate", 0.275)

    for i, row in enumerate(rows, 1):
        ticker = row.get("ticker","")
        name   = str(row.get("name", ticker))[:22]
        qty    = row.get("qty", row.get("quantity", 0))
        price  = row.get("price", row.get("current_price", 0))
        ccy    = row.get("currency","USD")
        div    = row.get("div_total", 0)
        kest   = row.get("kest_div", row.get("withholding", 0))
        qst    = row.get("foreign_tax", 0)
        note   = str(row.get("note", row.get("tax_note","")))[:24]

        dep_e  = to_eur(float(qty)*float(price), ccy)
        div_e  = to_eur(div, ccy)
        kest_e = to_eur(kest, ccy)
        qst_e  = to_eur(qst, ccy)
        tot_depot += dep_e; tot_div += div_e
        tot_kest  += kest_e; tot_qst += qst_e

        tbl_data.append([
            str(i),
            Paragraph(f"<b>{ticker}</b><br/><font size='7'>{name}</font>",
                      ParagraphStyle("tc", parent=styles["Normal"], fontSize=8, leading=10)),
            _n(qty,2), _n(price), ccy,
            eur(dep_e),
            eur(div_e)  if div_e  > 0 else "–",
            eur(kest_e) if kest_e > 0 else "–",
            eur(qst_e)  if qst_e  > 0 else "–",
            Paragraph(f"<font size='7'>{note}</font>",
                      ParagraphStyle("nt", parent=styles["Normal"], fontSize=7)),
        ])

    tbl_data.append(["", _Pb("TOTAL"), "", "", "",
                     _Pb(eur(tot_depot)), _Pb(eur(tot_div)),
                     _Pb(eur(tot_kest)), _Pb(eur(tot_qst)), ""])

    cw = [0.7*cm, 3.6*cm, 1.2*cm, 1.2*cm, 1.0*cm,
          2.2*cm, 2.0*cm, 1.8*cm, 1.8*cm, 2.0*cm]
    kt = Table(tbl_data, colWidths=cw, repeatRows=1)
    kt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  7.5),
        ("ALIGN",         (0,0),(-1,0),  "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("ALIGN",         (2,1),(-2,-1), "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGREY]),
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT),
        ("LINEABOVE",     (0,-1),(-1,-1), 1.0, DARK),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
    ]))
    story.append(kt)
    story.append(Spacer(1, 8))

    # E1kv Übertrag
    story.append(HRFlowable(width="100%", thickness=0.8, color=DARK))
    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>Übertrag in die Einkommensteuererklärung (E1kv)</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))

    verbleibend = max(0.0, tot_div * kest_rate - tot_qst)

    def _Pe(txt):
        return Paragraph(txt, ParagraphStyle("e", parent=styles["Normal"],
                                              fontSize=8, leading=10,
                                              textColor=colors.HexColor("#444")))

    def _Pb2(txt):
        return Paragraph(txt, ParagraphStyle("b2", parent=styles["Normal"],
                                              fontSize=8.5, leading=11))

    e1kv_rows = [
        ["KZ", "Bezeichnung", "Betrag EUR", "Erläuterung"],
        ["369", _Pb2("Kapitalerträge aus Wertpapieren (Dividenden)"),
         _Pb(eur(tot_div)), _Pe(f"Brutto-Dividenden ex-day-genau,\nSteuerjahr {year}")],
        ["–", _Pb2("KESt 27,5 % (durch Depot abgeführt)"),
         _Pb(eur(tot_kest)), _Pe("Bei österr. Depot:\nautomatisch abgeführt")],
        ["–", _Pb2("Ausländische Quellensteuer\n(anrechenbar auf KESt)"),
         _Pb(eur(tot_qst)), _Pe("Anrechenbar gemäss DBA\n(§ 48 BAO)")],
        ["–", _Pb2("Verbleibende KESt-Belastung (Schätzung)"),
         Paragraph(f"<b><font color='#c0392b'>{eur(verbleibend)}</font></b>",
                   ParagraphStyle("r", parent=styles["Normal"],
                                  fontSize=10, fontName="Helvetica-Bold")),
         _Pe("KESt nach Anrechnung\nausländ. Quellensteuer")],
        ["802", _Pb2("Depotwert per 31.12. (Vermögensaufstellung)"),
         _Pb(eur(tot_depot)), _Pe("Gesamtmarktwert aller\nPositionen in EUR")],
        ["–", _Pb2("Verlustausgleich (falls zutreffend)"),
         "____________________",
         _Pe("Automatisch durch Depotbank\n(§ 93 EStG)")],
    ]
    e1t = Table(e1kv_rows, colWidths=[1.5*cm, 6.8*cm, 3.2*cm, 5.5*cm])
    e1t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
        ("FONTSIZE",      (0,1),(-1,-1), 8.5),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("ALIGN",         (2,0),(2,-1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, YELL]),
    ]))
    story.append(e1t)
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey))
    story.append(Spacer(1, 3))
    story.append(Paragraph(
        f"⚠️  <b>Orientierungshilfe – kein offizielles FinanzOnline-Formular.</b> "
        f"KESt wird bei österreichischem Depot automatisch durch die Depotbank abgeführt. "
        f"Bitte verwenden Sie den originalen Depotauszug per 31.12.{year}. "
        f"Verlustausgleich erfolgt automatisch (§ 93 EStG). "
        f"Erstellt mit Stock Monitor v5.0.",
        ParagraphStyle("w", parent=styles["Normal"], fontSize=8, textColor=MID)
    ))
    doc.build(story)



def _export_uk_form_pdf(path, tax_data, nation_code, nation_name, year):
    """UK Self Assessment SA108 / Dividend Tax summary – Orientation aid."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    rows  = tax_data["rows"]
    rules = tax_data["rules"]

    fx_usd = _get_fx_rate_year_end("USD", "GBP", year)
    fx_eur = _get_fx_rate_year_end("EUR", "GBP", year)
    fx_chf = _get_fx_rate_year_end("CHF", "GBP", year)

    doc = SimpleDocTemplate(path, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    DARK  = colors.HexColor("#00247d")   # UK navy
    MID   = colors.HexColor("#cf142b")   # UK red
    LIGHT = colors.HexColor("#dde8ff")
    LGREY = colors.HexColor("#f5f5f5")
    YELL  = colors.HexColor("#fff9c4")
    RED   = colors.HexColor("#c0392b")

    def gbp(v):
        try:
            x = float(v)
            s = f"{abs(x):,.2f}"
            return f"-{s}" if x < 0 else s
        except: return "–"

    def _n(v, d=2):
        try: return f"{float(v):,.{d}f}"
        except: return "–"

    def to_gbp(val, ccy):
        try:
            v = float(val)
            if ccy == "GBP": return v
            if ccy == "EUR": return v * fx_eur
            if ccy == "CHF": return v * fx_chf
            return v * fx_usd
        except: return 0.0

    def _Pb(txt):
        return Paragraph(f"<b>{txt}</b>",
                         ParagraphStyle("b", parent=styles["Normal"],
                                        fontSize=8.5, fontName="Helvetica-Bold"))

    def _Pe(txt):
        return Paragraph(txt, ParagraphStyle("e", parent=styles["Normal"],
                                              fontSize=8, leading=10,
                                              textColor=colors.HexColor("#444")))

    # Header
    hdr = Table([[
        Paragraph("<b>UK Self Assessment – Dividends &amp; Capital Gains Summary</b>",
                  ParagraphStyle("h1", parent=styles["Normal"], fontSize=12,
                                 fontName="Helvetica-Bold", textColor=DARK)),
        Paragraph(f"Orientation Aid\nTax Year {year}/{str(year+1)[2:]}",
                  ParagraphStyle("r", parent=styles["Normal"], fontSize=8,
                                 alignment=2, textColor=colors.HexColor("#888")))
    ]], colWidths=[12*cm, 5*cm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), LIGHT),
        ("BOX",           (0,0),(-1,-1), 0.5, DARK),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    story.append(hdr)
    story.append(Paragraph(
        f"<font size='9' color='#00247d'>{nation_name} &nbsp;|&nbsp; "
        f"Tax Year 6 April {year} – 5 April {year+1} &nbsp;|&nbsp; "
        f"SA100 / SA108 / Dividend Tax</font>",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
                       spaceAfter=6, spaceBefore=3)
    ))

    # Personal details
    story.append(Paragraph("<b>Taxpayer Details</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))
    pt = Table([
        ["Full Name:",       "____________________________",
         "UTR / NI No.:",    "____________________________"],
        ["Address:",         "____________________________",
         "Tax Year:",        f"6 Apr {year} – 5 Apr {year+1}"],
    ], colWidths=[3.2*cm, 6.2*cm, 3.0*cm, 4.6*cm])
    pt.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8.5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LINEBELOW",     (1,0),(1,0),   0.4, colors.grey),
        ("LINEBELOW",     (3,0),(3,0),   0.4, colors.grey),
        ("LINEBELOW",     (1,1),(1,1),   0.4, colors.grey),
    ]))
    story.append(pt)
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=0.5, color=DARK))
    story.append(Paragraph(
        f"<font size='8'><b>FX Rates (approx. 5 Apr {year+1}, Yahoo Finance):</b> "
        f"1 USD = GBP {_n(fx_usd,4)} &nbsp;|&nbsp; "
        f"1 EUR = GBP {_n(fx_eur,4)} &nbsp;|&nbsp; "
        f"1 CHF = GBP {_n(fx_chf,4)}</font>",
        ParagraphStyle("fx", parent=styles["Normal"], fontSize=8,
                       spaceBefore=4, spaceAfter=8,
                       textColor=colors.HexColor("#444"))
    ))

    # Holdings table
    story.append(Paragraph("<b>Holdings &amp; Dividend Income – Individual Positions</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=3)))
    HDR = ["No.", "Security", "Qty", f"Price\n5 Apr", "CCY",
           "Market Value GBP", "Dividends GBP", "Unrealised G/L GBP", "Note"]
    tbl_data = [HDR]

    tot_mv = tot_div = tot_unr = 0.0

    for i, row in enumerate(rows, 1):
        ticker  = row.get("ticker", "")
        name    = str(row.get("name", ticker))[:22]
        qty     = row.get("qty", row.get("quantity", 0))
        price   = row.get("price", row.get("current_price", 0))
        ccy     = row.get("currency", "USD")
        div     = row.get("div_total", 0)
        unr     = row.get("unrealised", 0)
        note    = str(row.get("note", row.get("rate_note", "")))[:22]

        mv_g  = to_gbp(float(qty) * float(price), ccy)
        div_g = to_gbp(div, ccy)
        unr_g = to_gbp(unr, ccy)
        tot_mv  += mv_g
        tot_div += div_g
        tot_unr += unr_g

        tbl_data.append([
            str(i),
            Paragraph(f"<b>{ticker}</b><br/><font size='7'>{name}</font>",
                      ParagraphStyle("tc", parent=styles["Normal"], fontSize=8, leading=10)),
            _n(qty, 2), _n(price), ccy,
            gbp(mv_g),
            gbp(div_g) if div_g > 0 else "–",
            Paragraph(
                f"<font color='{'#27ae60' if unr_g >= 0 else '#c0392b'}'>{gbp(unr_g)}</font>",
                ParagraphStyle("u", parent=styles["Normal"], fontSize=8)
            ),
            Paragraph(f"<font size='7'>{note}</font>",
                      ParagraphStyle("nt", parent=styles["Normal"], fontSize=7)),
        ])

    tbl_data.append(["", _Pb("TOTAL"), "", "", "",
                     _Pb(gbp(tot_mv)), _Pb(gbp(tot_div)),
                     _Pb(gbp(tot_unr)), ""])

    cw = [0.7*cm, 3.8*cm, 1.2*cm, 1.3*cm, 0.9*cm,
          2.4*cm, 2.2*cm, 2.4*cm, 2.1*cm]
    kt = Table(tbl_data, colWidths=cw, repeatRows=1)
    kt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  7.5),
        ("ALIGN",         (0,0),(-1,0),  "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("ALIGN",         (2,1),(-2,-1), "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGREY]),
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT),
        ("LINEABOVE",     (0,-1),(-1,-1), 1.0, DARK),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
    ]))
    story.append(kt)
    story.append(Spacer(1, 8))

    # SA108 / Dividend Tax summary table
    story.append(HRFlowable(width="100%", thickness=0.8, color=DARK))
    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>Self Assessment Summary – Dividend Tax (SA100 Box / SA108)</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))

    allowance   = float(rules.get("dividend_allowance", 500))
    taxable_div = max(0.0, tot_div - allowance)
    div_tax_br  = taxable_div * rules.get("div_basic_rate", 0.0875)
    div_tax_hr  = taxable_div * rules.get("div_higher_rate", 0.3375)
    cgt_exempt  = float(rules.get("cgt_allowance", 3000))
    taxable_cgt = max(0.0, tot_unr - cgt_exempt) if tot_unr > 0 else 0.0
    cgt_basic   = taxable_cgt * rules.get("cgt_basic_rate", 0.10)
    cgt_higher  = taxable_cgt * rules.get("cgt_higher_rate", 0.20)

    sa_rows = [
        ["Box", "Description", "Amount GBP", "Notes"],
        ["SA100\nBox 3",
         _Pe("Total Dividend Income (gross)"),
         _Pb(gbp(tot_div)),
         _Pe(f"All dividends received\ntax year {year}/{str(year+1)[2:]}")],
        ["–",
         _Pe(f"Dividend Allowance (tax-free)"),
         _Pb(f"– {gbp(allowance)}"),
         _Pe("£500 p.a. (from April 2024)\n(§ Finance Act 2022)")],
        ["–",
         _Pe("Taxable Dividend Income"),
         Paragraph(f"<b><font color='#c0392b'>{gbp(taxable_div)}</font></b>",
                   ParagraphStyle("r", parent=styles["Normal"],
                                  fontSize=9, fontName="Helvetica-Bold")),
         _Pe("Basis for dividend tax")],
        ["–",
         _Pe("Dividend Tax – Basic Rate 8.75 %"),
         _Pb(gbp(div_tax_br)),
         _Pe("If total income within\nBasic Rate band")],
        ["–",
         _Pe("Dividend Tax – Higher Rate 33.75 %"),
         _Pb(gbp(div_tax_hr)),
         _Pe("If total income in\nHigher Rate band (>£50,270)")],
        ["SA108\nBox 6",
         _Pe("Total Unrealised Gains (all positions)"),
         _Pb(gbp(tot_unr)),
         _Pe("Only taxable upon disposal\n(realisation)")],
        ["–",
         _Pe(f"CGT Annual Exempt Amount"),
         _Pb(f"– {gbp(cgt_exempt)}"),
         _Pe("£3,000 p.a. (from April 2024)")],
        ["–",
         _Pe("Potentially Taxable Capital Gain"),
         Paragraph(f"<b><font color='#c0392b'>{gbp(taxable_cgt)}</font></b>",
                   ParagraphStyle("r", parent=styles["Normal"],
                                  fontSize=9, fontName="Helvetica-Bold")),
         _Pe("CGT Basic 10 % / Higher 20 %\n(upon realisation)")],
        ["–",
         _Pe("Estimated CGT (Basic Rate 10 %)"),
         _Pb(gbp(cgt_basic)),
         _Pe("If gains realised this year")],
        ["–",
         _Pe("Estimated CGT (Higher Rate 20 %)"),
         _Pb(gbp(cgt_higher)),
         _Pe("Higher/Additional Rate taxpayer")],
        ["–",
         _Pe("Total Portfolio Market Value (GBP)"),
         _Pb(gbp(tot_mv)),
         _Pe("For wealth / estate planning")],
    ]
    sat = Table(sa_rows, colWidths=[1.5*cm, 6.5*cm, 3.2*cm, 5.8*cm])
    sat.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
        ("FONTSIZE",      (0,1),(-1,-1), 8.5),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("ALIGN",         (2,0),(2,-1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, YELL]),
    ]))
    story.append(sat)
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey))
    story.append(Spacer(1, 3))
    story.append(Paragraph(
        f"⚠️  <b>Orientation aid – not an official HMRC form.</b> "
        f"Tax rates depend on total income. Basic Rate band: up to £50,270 total income. "
        f"Scotland has different dividend tax rates. "
        f"Please compare with your broker's Consolidated Tax Certificate and file via HMRC Self Assessment. "
        f"Capital Gains are only taxable on realisation (disposal), not on unrealised gains shown here. "
        f"Created with Stock Monitor.",
        ParagraphStyle("w", parent=styles["Normal"], fontSize=8, textColor=RED)
    ))
    doc.build(story)


def _export_us_form_pdf(path, tax_data, inv_code, inv_name, year,
                        state_code=None, state_name=None,
                        filing_code=None, filing_name=None):
    """US Schedule B / 1099-DIV / Form 1040 summary – Orientation aid. Letter format."""
    from reportlab.lib.pagesizes import LETTER
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch, cm

    rows  = tax_data["rows"]
    rules = tax_data["rules"]

    fx_eur = _get_fx_rate_year_end("EUR", "USD", year)
    fx_chf = _get_fx_rate_year_end("CHF", "USD", year)
    fx_gbp = _get_fx_rate_year_end("GBP", "USD", year)

    # Letter: 8.5" x 11" – US standard
    doc = SimpleDocTemplate(path, pagesize=LETTER,
                            rightMargin=1*inch, leftMargin=1*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story  = []

    DARK  = colors.HexColor("#002868")   # US navy
    MID   = colors.HexColor("#bf0a30")   # US red
    LIGHT = colors.HexColor("#dde8ff")
    LGREY = colors.HexColor("#f5f5f5")
    YELL  = colors.HexColor("#fff9c4")
    RED   = colors.HexColor("#c0392b")

    # Determine withholding rate from inv_code
    if inv_code in ("CH", "DE", "AT"):
        wh_rate = rules.get("dba_ch_rate", 0.15)
        wh_label = f"DBA {inv_code}-US 15 %"
    elif inv_code == "RES":
        wh_rate = rules.get("qualified_div_rates", [0, 0.15, 0.20])[1]
        wh_label = "Qualified Dividends 15 % (estimated)"
    else:
        wh_rate = rules.get("nra_withholding", 0.30)
        wh_label = "Standard NRA 30 %"

    def usd(v):
        try:
            x = float(v)
            s = f"{abs(x):,.2f}"
            return f"-{s}" if x < 0 else s
        except: return "–"

    def _n(v, d=2):
        try: return f"{float(v):,.{d}f}"
        except: return "–"

    def to_usd(val, ccy):
        try:
            v = float(val)
            if ccy == "USD": return v
            if ccy == "EUR": return v * fx_eur
            if ccy == "CHF": return v * fx_chf
            if ccy == "GBP": return v * fx_gbp
            return v
        except: return 0.0

    def _Pb(txt):
        return Paragraph(f"<b>{txt}</b>",
                         ParagraphStyle("b", parent=styles["Normal"],
                                        fontSize=8.5, fontName="Helvetica-Bold"))

    def _Pe(txt):
        return Paragraph(txt, ParagraphStyle("e", parent=styles["Normal"],
                                              fontSize=8, leading=10,
                                              textColor=colors.HexColor("#444")))

    # Header
    hdr = Table([[
        Paragraph("<b>US Tax Summary – Schedule B / 1099-DIV / Form 1040</b>",
                  ParagraphStyle("h1", parent=styles["Normal"], fontSize=12,
                                 fontName="Helvetica-Bold", textColor=DARK)),
        Paragraph(f"Orientation Aid\nTax Year {year}",
                  ParagraphStyle("r", parent=styles["Normal"], fontSize=8,
                                 alignment=2, textColor=colors.HexColor("#888")))
    ]], colWidths=[12*cm, 5*cm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), LIGHT),
        ("BOX",           (0,0),(-1,-1), 0.5, DARK),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    story.append(hdr)
    _state_hint   = (f" &nbsp;|&nbsp; State: {state_name.split('–')[0].strip()}"
                     if inv_code == "RES" and state_name else "")
    _filing_hint  = (f" &nbsp;|&nbsp; Filing: {filing_name}"
                     if inv_code == "RES" and filing_name else "")
    story.append(Paragraph(
        f"<font size='9' color='#002868'>Investor: {inv_name} &nbsp;|&nbsp; "
        f"Tax Year {year} &nbsp;|&nbsp; Withholding: {wh_label}"
        f"{_state_hint}{_filing_hint}</font>",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
                       spaceAfter=6, spaceBefore=3)
    ))

    # Personal details
    story.append(Paragraph("<b>Taxpayer Information</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))
    pt = Table([
        ["Name:",          "____________________________",
         "SSN / ITIN / TIN:", "____________________________"],
        ["Address:",        "____________________________",
         "Tax Year:",       str(year)],
    ], colWidths=[1.2*inch, 2.3*inch, 1.4*inch, 1.6*inch])
    pt.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8.5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LINEBELOW",     (1,0),(1,0),   0.4, colors.grey),
        ("LINEBELOW",     (3,0),(3,0),   0.4, colors.grey),
        ("LINEBELOW",     (1,1),(1,1),   0.4, colors.grey),
    ]))
    story.append(pt)
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=0.5, color=DARK))
    story.append(Paragraph(
        f"<font size='8'><b>FX Rates (31 Dec {year}, Yahoo Finance):</b> "
        f"1 EUR = USD {_n(fx_eur,4)} &nbsp;|&nbsp; "
        f"1 CHF = USD {_n(fx_chf,4)} &nbsp;|&nbsp; "
        f"1 GBP = USD {_n(fx_gbp,4)}</font>",
        ParagraphStyle("fx", parent=styles["Normal"], fontSize=8,
                       spaceBefore=4, spaceAfter=8,
                       textColor=colors.HexColor("#444"))
    ))

    # Wash Sale warning box
    story.append(Table([[
        Paragraph(
            "<b>⚠️  Wash Sale Rule (IRC § 1091) – NOT calculated automatically</b><br/>"
            "If you sold a security at a loss and repurchased the same or substantially "
            "identical security within 30 days before or after the sale, the loss is "
            "<b>disallowed</b> for tax purposes. The disallowed loss is added to the "
            "cost basis of the replacement shares.<br/>"
            "<b>Action required:</b> Check your broker's 1099-B for 'Wash Sale Loss Disallowed' entries.",
            ParagraphStyle("ws", parent=styles["Normal"], fontSize=8, leading=11,
                           textColor=colors.HexColor("#7d4800"))
        )
    ]], colWidths=[6.5*inch]))
    story[-1].setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#fff3cd")),
        ("BOX",           (0,0),(-1,-1), 1.0, colors.HexColor("#e6a817")),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 10),
        ("TOPPADDING",    (0,0),(-1,-1), 7),
        ("BOTTOMPADDING", (0,0),(-1,-1), 7),
    ]))
    story.append(Spacer(1, 8))

    # Holdings table
    story.append(Paragraph("<b>Schedule B – Dividend &amp; Interest Income – Individual Positions</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=3)))
    HDR = ["No.", "Security", "Qty", f"Price\n31 Dec", "CCY",
           "Market Value USD", "Dividends USD", f"Withholding\n{wh_label[:16]}", "Note"]
    tbl_data = [HDR]

    tot_mv = tot_div = tot_wh = 0.0

    for i, row in enumerate(rows, 1):
        ticker = row.get("ticker", "")
        name   = str(row.get("name", ticker))[:22]
        qty    = row.get("qty", row.get("quantity", 0))
        price  = row.get("price", row.get("current_price", 0))
        ccy    = row.get("currency", "USD")
        div    = row.get("div_total", 0)
        wh     = row.get("tax", row.get("withholding", div * wh_rate))
        note   = str(row.get("rate_note", row.get("note", "")))[:22]

        mv_u  = to_usd(float(qty) * float(price), ccy)
        div_u = to_usd(div, ccy)
        wh_u  = to_usd(wh, ccy)
        tot_mv  += mv_u
        tot_div += div_u
        tot_wh  += wh_u

        tbl_data.append([
            str(i),
            Paragraph(f"<b>{ticker}</b><br/><font size='7'>{name}</font>",
                      ParagraphStyle("tc", parent=styles["Normal"], fontSize=8, leading=10)),
            _n(qty, 2), _n(price), ccy,
            usd(mv_u),
            usd(div_u) if div_u > 0 else "–",
            usd(wh_u)  if wh_u  > 0 else "–",
            Paragraph(f"<font size='7'>{note}</font>",
                      ParagraphStyle("nt", parent=styles["Normal"], fontSize=7)),
        ])

    tbl_data.append(["", _Pb("TOTAL"), "", "", "",
                     _Pb(usd(tot_mv)), _Pb(usd(tot_div)), _Pb(usd(tot_wh)), ""])

    # Letter 6.5" usable width: 0.3+1.6+0.5+0.55+0.4+1.0+0.95+1.0+0.7 = 7.0 → scaled
    cw = [0.3*inch, 1.6*inch, 0.5*inch, 0.55*inch, 0.4*inch,
          1.0*inch, 0.95*inch, 1.0*inch, 0.7*inch]
    kt = Table(tbl_data, colWidths=cw, repeatRows=1)
    kt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  7.5),
        ("ALIGN",         (0,0),(-1,0),  "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("ALIGN",         (2,1),(-2,-1), "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGREY]),
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT),
        ("LINEABOVE",     (0,-1),(-1,-1), 1.0, DARK),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
    ]))
    story.append(kt)
    story.append(Spacer(1, 8))

    # Form 1040 / Schedule B summary
    story.append(HRFlowable(width="100%", thickness=0.8, color=DARK))
    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>Form 1040 / Schedule B Summary – Transfer Guidance</b>",
                           ParagraphStyle("h2", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=DARK,
                                          spaceAfter=4)))

    net_div = tot_div - tot_wh

    if inv_code == "RES":
        # US Resident: qualified dividend logic – Schwellen je Filing Status (2024)
        qual_rate_low  = rules.get("qualified_div_rates", [0, 0.15, 0.20])[0]
        qual_rate_mid  = rules.get("qualified_div_rates", [0, 0.15, 0.20])[1]
        qual_rate_high = rules.get("qualified_div_rates", [0, 0.15, 0.20])[2]

        # Einkommensschwellen 2024 je Filing Status für Qualified Dividends / LTCG
        _filing_thresholds = {
            # (0%-Grenze, 15%-Grenze)  →  über 15%-Grenze: 20%
            "SINGLE": (47_025,   518_900),
            "MFJ":    (94_050,   583_750),   # Married Filing Jointly
            "MFS":    (47_025,   291_850),   # Married Filing Separately
            "HOH":    (63_000,   551_350),   # Head of Household
        }
        _fs = filing_code or "SINGLE"
        _thresh_0pct, _thresh_15pct = _filing_thresholds.get(_fs, (47_025, 518_900))
        _filing_disp = filing_name or "Single"

        # NIIT: 3.8 % auf Kapitalerträge ab $200k (Single) / $250k (MFJ)
        _niit_thresh = 250_000 if _fs == "MFJ" else 200_000
        _niit_rate   = 0.038
        _niit_est    = tot_div * _niit_rate  # Schätzung wenn Einkommen über Schwelle

        # State tax estimate
        _state_rates = {
            # No state income tax
            "NONE": 0.0, "NH": 0.03, "TN": 0.0,
            "WY_note": 0.0, "WA_note": 0.0, "SD_note": 0.0,
            "AK_note": 0.0, "FL_note": 0.0, "TX_note": 0.0, "NV_note": 0.0,
            # High
            "CA": 0.093,  "DC": 0.085,  "NJ": 0.0637, "OR": 0.0875,
            "MN": 0.0785, "NY": 0.0685, "VT": 0.066,  "HI": 0.0825,
            "CT": 0.05,   "WI": 0.0765, "SC": 0.065,  "ME": 0.0715,
            "ID": 0.0569, "MT": 0.059,  "AR": 0.049,  "NM": 0.059,
            "DE": 0.066,  "NE": 0.0584, "WV": 0.065,  "RI": 0.0599,
            "MD": 0.0575,
            # Medium
            "VA": 0.0575, "MA": 0.05,   "IL": 0.0495, "KY": 0.04,
            "LA": 0.0425, "MO": 0.0495, "AL": 0.05,   "OH": 0.0399,
            "KS": 0.057,  "OK": 0.0475, "MS": 0.047,  "IA": 0.06,
            "GA": 0.0549, "NC": 0.0475, "MI": 0.0405, "PA": 0.0307,
            # Low
            "UT": 0.0465, "IN": 0.0305, "CO": 0.044,  "AZ": 0.025,
            "ND": 0.025,
            # Fallback
            "OTHER": 0.05,
        }
        st_rate = _state_rates.get(state_code or "NONE", 0.0)
        st_tax  = tot_div * st_rate
        st_disp = state_name.split("–")[0].strip() if state_name else "N/A"
        f1040_rows = [
            ["Line", "Description", "Amount USD", "Notes"],
            ["Sch B\nLine 1a", _Pe("Total Ordinary Dividends (1099-DIV Box 1a)"),
             _Pb(usd(tot_div)), _Pe(f"All dividends, tax year {year}")],
            ["Sch B\nLine 1b", _Pe("Qualified Dividends (1099-DIV Box 1b)"),
             _Pb(usd(tot_div)),
             _Pe("Assumed 100 % qualified\n(holding period > 60 days)")],
            ["–", _Pe("Federal Tax withheld (1099-DIV Box 4)"),
             _Pb(usd(tot_wh)), _Pe("Credit against tax liability")],
            ["–", _Pe(f"Filing Status: {_filing_disp}"),
             _Pb("–"),
             _Pe(f"0 % up to ${_thresh_0pct:,}\n"
                 f"15 % up to ${_thresh_15pct:,}\n"
                 f"20 % above ${_thresh_15pct:,}")],
            ["–", _Pe(f"Qualified Div Tax – Rate 0 % (income ≤ ${_thresh_0pct:,})"),
             _Pb("0.00"), _Pe(f"Tax-free bracket\n({_filing_disp}, 2024)")],
            ["–", _Pe(f"Qualified Div Tax – Rate 15 % (income ≤ ${_thresh_15pct:,})"),
             _Pb(usd(tot_div * qual_rate_mid)),
             _Pe(f"Most common rate\n({_filing_disp})")],
            ["–", _Pe(f"Qualified Div Tax – Rate 20 % (income > ${_thresh_15pct:,})"),
             _Pb(usd(tot_div * qual_rate_high)),
             _Pe(f"High-income bracket\n({_filing_disp})")],
            ["–", _Pe("Net Dividends after estimated federal withholding"),
             Paragraph(f"<b><font color='#002868'>{usd(net_div)}</font></b>",
                       ParagraphStyle("r", parent=styles["Normal"],
                                      fontSize=9, fontName="Helvetica-Bold")),
             _Pe("After broker withholding")],
            ["State\nTax", _Pe(f"State Income Tax – {st_disp} (~{st_rate*100:.2f} %)"),
             Paragraph(f"<b><font color='#bf0a30'>{usd(st_tax)}</font></b>",
                       ParagraphStyle("r", parent=styles["Normal"],
                                      fontSize=9, fontName="Helvetica-Bold")),
             _Pe("Estimated – verify with\nstate tax return")],
            ["NIIT\n3.8 %", _Pe(f"Net Investment Income Tax (~${_niit_thresh:,} threshold)"),
             _Pb(usd(_niit_est)),
             _Pe(f"Only if total income\nexceeds ${_niit_thresh:,}\n({_filing_disp})")],
            ["–", _Pe("Total estimated burden (Federal 15 % + State + NIIT)"),
             Paragraph(f"<b><font color='#bf0a30'>"
                       f"{usd(tot_div * qual_rate_mid + st_tax + _niit_est)}"
                       f"</font></b>",
                       ParagraphStyle("r", parent=styles["Normal"],
                                      fontSize=10, fontName="Helvetica-Bold")),
             _Pe("Worst-case estimate\n(verify with tax advisor)")],
            ["–", _Pe("Total Portfolio Market Value (USD)"),
             _Pb(usd(tot_mv)), _Pe("Dec 31 valuation")],
        ]
    else:
        # Non-US person
        f1040_rows = [
            ["Form", "Description", "Amount USD", "Notes"],
            ["1042-S\nBox 2", _Pe("Gross Income (Dividends)"),
             _Pb(usd(tot_div)), _Pe(f"All dividends, tax year {year}")],
            ["1042-S\nBox 7", _Pe(f"Withholding Tax ({wh_label})"),
             _Pb(usd(tot_wh)), _Pe("Withheld at source by broker")],
            ["–", _Pe("Net Dividend Income (after withholding)"),
             Paragraph(f"<b><font color='#002868'>{usd(net_div)}</font></b>",
                       ParagraphStyle("r", parent=styles["Normal"],
                                      fontSize=9, fontName="Helvetica-Bold")),
             _Pe("Paid out to investor")],
            ["–", _Pe("DBA Treaty / Rate Applied"),
             _Pb(wh_label), _Pe("Verify with broker 1042-S")],
            ["–", _Pe("Wash Sale Rule (IRC § 1091)"),
             Paragraph("<b><font color='#c0392b'>NOT calculated</font></b>",
                       ParagraphStyle("r", parent=styles["Normal"],
                                      fontSize=8.5, fontName="Helvetica-Bold")),
             _Pe("Check 1099-B for\ndisallowed losses")],
            ["–", _Pe("FBAR / FATCA reporting threshold"),
             _Pb("$10,000 / $50,000"),
             _Pe("Foreign Bank Account Report\n(FinCEN 114) if applicable")],
            ["–", _Pe("Total Portfolio Market Value (USD)"),
             _Pb(usd(tot_mv)), _Pe("Dec 31 valuation")],
        ]

    f1t = Table(f1040_rows, colWidths=[0.6*inch, 2.8*inch, 1.3*inch, 1.8*inch])
    f1t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("BOX",           (0,0),(-1,-1), 0.8, DARK),
        ("FONTSIZE",      (0,1),(-1,-1), 8.5),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("ALIGN",         (2,0),(2,-1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, YELL]),
    ]))
    story.append(f1t)
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey))
    story.append(Spacer(1, 3))
    story.append(Paragraph(
        f"⚠️  <b>Orientation aid – not an official IRS form.</b> "
        f"Non-US persons: compare with broker 1042-S statement. "
        f"US residents: compare with 1099-DIV and file Schedule B with Form 1040. "
        f"Wash Sale Rule (IRC § 1091, 30-day window) is NOT calculated – check your 1099-B. "
        f"FBAR filing required if foreign accounts exceed $10,000 at any point in {year}. "
        f"Rates shown are estimates; actual liability depends on total income, filing status, and state taxes. "
        f"Created with Stock Monitor.",
        ParagraphStyle("w", parent=styles["Normal"], fontSize=8, textColor=RED)
    ))
    doc.build(story)


# ── Export-Funktionen ─────────────────────────────────────────────────────────

def _export_pdf(path, tax_data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    rules   = tax_data["rules"]
    rows    = tax_data["rows"]
    summary = tax_data["summary"]
    country = tax_data["country"]
    cur     = summary.get("currency", "–")

    doc  = SimpleDocTemplate(path, pagesize=landscape(A4),
                              rightMargin=1.5*cm, leftMargin=1.5*cm,
                              topMargin=1.5*cm,  bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, spaceAfter=6)
    note_style  = ParagraphStyle("note",  parent=styles["Normal"], fontSize=8,
                                 textColor=colors.HexColor("#555555"), spaceAfter=4)

    story.append(Paragraph(
        f"{rules['name']} – Steuerauszug {rules['year']}", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(rules["disclaimer"], note_style))
    story.append(Spacer(1, 8))

    # Tabellen-Header
    _tc = TRT("tax_col_ticker"); _tn = TRT("tax_col_name")
    _tq = TRT("tax_col_qty");   _tp = TRT("tax_col_price")
    _tm = TRT("tax_col_market_val"); _td = TRT("tax_col_div")
    if country == "AT":
        t_headers = [_tc, _tn, _tq, _tp, _tm, _td,
                     TRT("tax_col_kest_div_short"), TRT("tax_col_foreign_short"), TRT("tax_col_note")]
    elif country == "CH":
        t_headers = [_tc, _tn, _tq, _tp, _tm, _td,
                     TRT("tax_col_vst_short"), TRT("tax_col_foreign_short"), TRT("tax_col_note")]
    elif country == "DE":
        t_headers = [_tc, _tn, _tq, _tp, _tm, _td, TRT("tax_col_unrealised_short")]
    elif country == "UK":
        t_headers = [_tc, _tn, _tq, _tp, _tm, _td, TRT("tax_col_unrealised_short")]
    else:
        t_headers = [_tc, _tn, _tq, _tp, _tm, _td,
                     TRT("tax_col_withholding"), TRT("tax_col_rate")]

    def _f(v, d=2):
        try: return _fmt(float(v), d)
        except: return str(v)

    tbl_data = [t_headers]
    for row in rows:
        if country == "AT":
            tbl_data.append([
                row["ticker"], row["name"][:22],
                _f(row["qty"], 0), _f(row["price"]), _f(row["market_val"]),
                _f(row["div_total"]), _f(row["kest_div"]), _f(row["foreign_tax"]),
                row["tax_note"][:20],
            ])
        elif country == "CH":
            tbl_data.append([
                row["ticker"], row["name"][:22],
                _f(row["qty"], 0), _f(row["price"]), _f(row["market_val"]),
                _f(row["div_total"]), _f(row["vst"]), _f(row["foreign_tax"]),
                row["tax_note"][:20],
            ])
        elif country in ("DE", "UK"):
            tbl_data.append([
                row["ticker"], row["name"][:22],
                _f(row["qty"], 0), _f(row["price"]), _f(row["market_val"]),
                _f(row["div_total"]), _f(row.get("unrealised", 0)),
            ])
        else:
            tbl_data.append([
                row["ticker"], row["name"][:22],
                _f(row["qty"], 0), _f(row["price"]), _f(row["market_val"]),
                _f(row["div_total"]), _f(row.get("tax", 0)),
                row.get("rate_note", ""),
            ])

    col_w = [1.5*cm, 4.5*cm, 1.3*cm, 2.0*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3.0*cm]
    t = Table(tbl_data, colWidths=col_w[:len(t_headers)], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("FONTSIZE",    (0,1), (-1,-1), 7),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.HexColor("#f0f4ff"), colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("ALIGN",       (2,0), (-1,-1), "RIGHT"),
        ("ALIGN",       (0,0), (1,-1),  "LEFT"),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))

    # Zusammenfassung
    sum_style = ParagraphStyle("sum", parent=styles["Normal"], fontSize=9, spaceAfter=2)
    story.append(Paragraph(f"<b>{TRT('tax_lbl_summary')}</b>", sum_style))

    if country == "AT":
        sum_keys = [
            (TRT("tax_sum_market_val").rstrip(":"),    "total_market_value"),
            (TRT("tax_sum_gross_div").rstrip(":"),     "total_dividends"),
            (TRT("tax_sum_kest_div").rstrip(":"),      "total_kest_div"),
            (TRT("tax_sum_foreign_tax").rstrip(":"),   "total_foreign_tax"),
            (TRT("tax_sum_kest_gain").rstrip(":"),     "total_kest_gain"),
            (TRT("tax_sum_total_tax").rstrip(":"),     "total_tax"),
            (TRT("tax_sum_net_div_kest").rstrip(":"),  "net_dividends"),
        ]
    else:
        sum_keys = [
            (TRT("tax_sum_market_val").rstrip(":"),    "total_market_value"),
            (TRT("tax_sum_gross_div").rstrip(":"),     "total_dividends"),
            (TRT("tax_sum_total_tax").rstrip(":"),     "total_tax"),
            (TRT("tax_sum_net_div_kest").rstrip(":"),  "net_dividends"),
        ]
    for label, key in sum_keys:
        val = summary.get(key, 0)
        story.append(Paragraph(
            f"{label}: <b>{_f(val)}  {cur}</b>", sum_style))

    story.append(Spacer(1, 8))
    story.append(Paragraph(_disclaimer_global(), note_style))
    doc.build(story)


def _export_xlsx(path, tax_data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    rules   = tax_data["rules"]
    rows    = tax_data["rows"]
    summary = tax_data["summary"]
    country = tax_data["country"]
    cur     = summary.get("currency", "–")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tax {country} {rules['year']}"

    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titel
    ws.append([f"{rules['name']} – {TRT('tax_lbl_tax_statement')} {rules['year']}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([rules["disclaimer"]])
    ws.append([])

    _tc = TRT("tax_col_ticker"); _tn = TRT("tax_col_name")
    _tq = TRT("tax_col_qty");   _tp = TRT("tax_col_price")
    _tm = TRT("tax_col_market_val"); _td = TRT("tax_col_div")
    if country == "AT":
        headers = [_tc, _tn, _tq, _tp, _tm,
                   f"{_td} ({cur})", f"{TRT('tax_col_kest_div_short')} ({cur})",
                   f"{TRT('tax_col_foreign_short')} ({cur})", TRT("tax_col_note")]
    elif country == "CH":
        headers = [_tc, _tn, _tq, _tp, _tm,
                   f"{_td} ({cur})", f"{TRT('tax_col_vst_short')} ({cur})",
                   f"{TRT('tax_col_foreign_short')} ({cur})", TRT("tax_col_note")]
    elif country in ("DE", "UK"):
        headers = [_tc, _tn, _tq, _tp, _tm,
                   f"{_td} ({cur})", f"{TRT('tax_col_unrealised_short')} ({cur})"]
    else:
        headers = [_tc, _tn, _tq, _tp, _tm,
                   f"{_td} ({cur})", f"{TRT('tax_col_withholding')} ({cur})", TRT("tax_col_rate")]

    hdr_row = ws.max_row + 1
    ws.append(headers)
    for cell in ws[hdr_row]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.border = border; cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows):
        if country == "AT":
            r = [row["ticker"], row["name"],
                 row["qty"], row["price"], row["market_val"],
                 row["div_total"], row["kest_div"], row["foreign_tax"], row["tax_note"]]
        elif country == "CH":
            r = [row["ticker"], row["name"],
                 row["qty"], row["price"], row["market_val"],
                 row["div_total"], row["vst"], row["foreign_tax"], row["tax_note"]]
        elif country in ("DE", "UK"):
            r = [row["ticker"], row["name"],
                 row["qty"], row["price"], row["market_val"],
                 row["div_total"], row.get("unrealised", 0)]
        else:
            r = [row["ticker"], row["name"],
                 row["qty"], row["price"], row["market_val"],
                 row["div_total"], row.get("tax", 0), row.get("rate_note", "")]
        ws.append(r)
        data_row = ws.max_row
        if i % 2 == 0:
            for cell in ws[data_row]:
                cell.fill = alt_fill
        for cell in ws[data_row]:
            cell.border = border
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    ws.append([])
    ws.append([TRT("tax_lbl_summary")])
    ws["A" + str(ws.max_row)].font = Font(bold=True, size=11)

    if country == "AT":
        sum_rows = [
            (f"{TRT('tax_sum_market_val').rstrip(':')} ({cur})",           "total_market_value"),
            (f"{TRT('tax_sum_gross_div').rstrip(':')} ({cur})",            "total_dividends"),
            (f"{TRT('tax_sum_kest_div').rstrip(':')} ({cur})",             "total_kest_div"),
            (f"{TRT('tax_sum_foreign_tax').rstrip(':')} ({cur})",          "total_foreign_tax"),
            (f"{TRT('tax_sum_kest_gain').rstrip(':')} ({cur})",            "total_kest_gain"),
            (f"{TRT('tax_sum_total_tax').rstrip(':')} ({cur})",            "total_tax"),
            (f"{TRT('tax_sum_net_div_kest').rstrip(':')} ({cur})",         "net_dividends"),
        ]
    else:
        sum_rows = [
            (f"{TRT('tax_sum_market_val').rstrip(':')} ({cur})",  "total_market_value"),
            (f"{TRT('tax_sum_gross_div').rstrip(':')} ({cur})",   "total_dividends"),
            (f"{TRT('tax_sum_total_tax').rstrip(':')} ({cur})",   "total_tax"),
            (f"{TRT('tax_sum_net_div_kest').rstrip(':')} ({cur})", "net_dividends"),
        ]
    for label, key in sum_rows:
        ws.append([label, summary.get(key, 0)])
        r = ws.max_row
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"].number_format = "#,##0.00"

    ws.append([])
    ws.append([_disclaimer_global()])
    ws[f"A{ws.max_row}"].font = Font(italic=True, size=8, color="777777")

    # Spaltenbreite
    for col in ws.columns:
        max_l = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_l + 3, 40)

    wb.save(path)


def _export_ods(path, tax_data):
    try:
        import odf
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table        import Table, TableRow, TableCell
        from odf.text         import P
        from odf.style        import Style, TextProperties, TableCellProperties
        import odf.style
    except ImportError:
        raise ImportError(
            "Das Paket 'odfpy' ist nicht installiert.\n"
            "pip install odfpy --break-system-packages"
        )

    rules   = tax_data["rules"]
    rows    = tax_data["rows"]
    summary = tax_data["summary"]
    country = tax_data["country"]
    cur     = summary.get("currency", "–")

    doc = OpenDocumentSpreadsheet()
    sheet = Table(name=f"Tax_{country}_{rules['year']}")
    doc.spreadsheet.addElement(sheet)

    def _row(*vals):
        tr = TableRow()
        for v in vals:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        sheet.addElement(tr)

    _row(f"{rules['name']} – {TRT('tax_lbl_tax_statement')} {rules['year']}")
    _row(rules["disclaimer"])
    _row()

    _tc = TRT("tax_col_ticker"); _tn = TRT("tax_col_name")
    _tq = TRT("tax_col_qty");   _tp = TRT("tax_col_price")
    _tm = TRT("tax_col_market_val"); _td = TRT("tax_col_div")
    if country == "AT":
        _row(_tc, _tn, _tq, _tp, _tm, _td,
             TRT("tax_col_kest_div_short"), TRT("tax_col_foreign_short"), TRT("tax_col_note"))
        for row in rows:
            _row(row["ticker"],row["name"],row["qty"],row["price"],
                 round(row["market_val"],2),round(row["div_total"],2),
                 round(row["kest_div"],2),round(row["foreign_tax"],2),row["tax_note"])
    elif country == "CH":
        _row(_tc, _tn, _tq, _tp, _tm, _td,
             TRT("tax_col_vst_short"), TRT("tax_col_foreign_short"), TRT("tax_col_note"))
        for row in rows:
            _row(row["ticker"],row["name"],row["qty"],row["price"],
                 round(row["market_val"],2),round(row["div_total"],2),
                 round(row["vst"],2),round(row["foreign_tax"],2),row["tax_note"])
    elif country in ("DE","UK"):
        _row(_tc, _tn, _tq, _tp, _tm, _td, TRT("tax_col_unrealised_short"))
        for row in rows:
            _row(row["ticker"],row["name"],row["qty"],row["price"],
                 round(row["market_val"],2),round(row["div_total"],2),
                 round(row.get("unrealised",0),2))
    else:
        _row(_tc, _tn, _tq, _tp, _tm, _td, TRT("tax_col_withholding"), TRT("tax_col_rate"))
        for row in rows:
            _row(row["ticker"],row["name"],row["qty"],row["price"],
                 round(row["market_val"],2),round(row["div_total"],2),
                 round(row.get("tax",0),2),row.get("rate_note",""))

    _row()
    _row(TRT("tax_lbl_summary"))
    if country == "AT":
        ods_sum_rows = [
            (f"{TRT('tax_sum_market_val').rstrip(':')} ({cur})",          "total_market_value"),
            (f"{TRT('tax_sum_gross_div').rstrip(':')} ({cur})",            "total_dividends"),
            (f"{TRT('tax_sum_kest_div').rstrip(':')} ({cur})",             "total_kest_div"),
            (f"{TRT('tax_sum_foreign_tax').rstrip(':')} ({cur})",          "total_foreign_tax"),
            (f"{TRT('tax_sum_kest_gain').rstrip(':')} ({cur})",            "total_kest_gain"),
            (f"{TRT('tax_sum_total_tax').rstrip(':')} ({cur})",            "total_tax"),
            (f"{TRT('tax_sum_net_div_kest').rstrip(':')} ({cur})",         "net_dividends"),
        ]
    else:
        ods_sum_rows = [
            (f"{TRT('tax_sum_market_val').rstrip(':')} ({cur})",  "total_market_value"),
            (f"{TRT('tax_sum_gross_div').rstrip(':')} ({cur})",   "total_dividends"),
            (f"{TRT('tax_sum_total_tax').rstrip(':')} ({cur})",   "total_tax"),
            (f"{TRT('tax_sum_net_div_kest').rstrip(':')} ({cur})", "net_dividends"),
        ]
    for label, key in ods_sum_rows:
        _row(label, round(summary.get(key, 0), 2))
    _row()
    _row(_disclaimer_global())

    doc.save(path)
